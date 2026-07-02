"""
pgr_viewer_pro_v6.py
====================
Professional PGR Survey Viewer — Eng. Samir Abozahra Edition (v6.3 - UltraSpeed Performance)
─────────────────────────────────────────────────────────────────
NEW IN v6.3 (ULTRASPEED UPDATE):
• ⚡ Multi-threaded background loading for near-instant file indexing
• ⚡ Background frame prefetching during playback for stutter-free video
• ⚡ Adaptive resolution scaling: Ultra-fast UI panning/playback, restoring to Lanczos on pause
• ⚡ Unified cache engine utilized for all viewing modes
NEW IN v6.2:
• ✅ GR view panel zoom in/out (mouse wheel, +/- keys, toolbar +/- buttons)
• ✅ Click-and-drag panning while zoomed in, with cursor-anchored zoom
• ✅ High-quality LANCZOS resampling for crisp display at every zoom level
NEW IN v6.1:
• ✅ Checkbox dropdown for selecting specific PGR files to merge
─────────────────────────────────────────────────────────────────
"""

import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import os
import glob
import math
import mmap
import zipfile
import xml.etree.ElementTree as ET
import numpy as np
from PIL import Image, ImageTk, ImageFile
import struct
import threading
import csv
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed
from collections import OrderedDict
import weakref
import gc
import time
import sys
import io
import re
import json

ImageFile.LOAD_TRUNCATED_IMAGES = True

try:
    import imagecodecs
    HAVE_CODECS = True
except ImportError:
    HAVE_CODECS = False
    print("WARNING: imagecodecs not installed. Install with: pip install imagecodecs")

try:
    import pandas as pd
    HAVE_PANDAS = True
except ImportError:
    HAVE_PANDAS = False

try:
    import tkintermapview
    HAVE_MAP = True
except ImportError:
    HAVE_MAP = False
    print("WARNING: tkintermapview not installed. Install with: pip install tkintermapview")

try:
    import cv2
    HAVE_CV2 = True
except ImportError:
    HAVE_CV2 = False
    print("WARNING: opencv-python not installed. Install with: pip install opencv-python")

# ─────────────────────────────────────────────────────────────
# PGR binary format constants
# ─────────────────────────────────────────────────────────────
PGR_MAGIC               = b"PGRLADYBUGSTREAM"
CAFEBABE_NEEDLE         = b"\xca\xfe\xba\xbe"
PGR_CAFEBABE_VALUE      = 0xCAFEBABE
PGR_NUM_CAMERAS         = 6
PGR_PLANES_PER_CAMERA   = 4
PGR_MAX_SUBIMAGES       = PGR_NUM_CAMERAS * PGR_PLANES_PER_CAMERA   # 24
PGR_FRAME_HEADER_SIZE   = 1024
PGR_IMAGE_TABLE_OFFSET  = PGR_FRAME_HEADER_SIZE - PGR_MAX_SUBIMAGES * 8  # 832
PGR_MIN_SUBIMAGE_BYTES  = 1024

CAMERA_NAMES = {
    0: "Camera 0 – Front Left",
    1: "Camera 1 – Front Right",
    2: "Camera 2 – Right",
    3: "Camera 3 – Rear",
    4: "Camera 4 – Left",
    5: "Camera 5 – Top",
}

DUAL_VIEW_OPTIONS = {
    "Single Camera": "single",
    "Dual Panoramic (Cam 2+3)": "dual",
    "All 6 Cameras Grid": "grid",
}

MAP_TILES = {
    "Satellite (Google)": "https://mt0.google.com/vt/lyrs=s&hl=en&x={x}&y={y}&z={z}&s=Ga",
    "Road (Google)":      "https://mt0.google.com/vt/lyrs=m&hl=en&x={x}&y={y}&z={z}&s=Ga",
    "Hybrid (Google)":    "https://mt0.google.com/vt/lyrs=y&hl=en&x={x}&y={y}&z={z}&s=Ga",
}

# ── Performance tuning parameters ──────────────────────────────────
CAR_MARKER_COLOR = "#ff4444"
ROUTE_COLOR      = "#00d4ff"
ACCENT           = "#00d4ff"
ACCENT2          = "#7c4dff"
BG_DARK          = "#0b0e14"
BG_PANEL         = "#12161f"
BG_CARD          = "#1a1f2e"
BG_TOOLBAR       = "#0f1320"
FG_MAIN          = "#dde3f0"
FG_DIM           = "#6a7490"
FG_BRIGHT        = "#ffffff"
BTN_PRIMARY      = "#1756b8"
BTN_HOVER        = "#1e6fd6"
BTN_DARK         = "#0d1a35"
SUCCESS          = "#00c896"
WARNING          = "#f5a623"
DANGER           = "#ff4757"
MERGED_COLOR     = "#7c4dff"

GPS_MAX_SEGMENT_M = 50.0
STRIP_PLACEMARK_NAMES = {"trajectory_path - 0001", "trajectory_path-0001"}

# Matches "Run 0001", "Run0001", "run_0001", etc. -> captures the 4-digit run number
RUN_PLACEMARK_PATTERN = re.compile(r"run[\s_-]*0*(\d+)", re.IGNORECASE)
# Matches the run number embedded in a PGR filename, e.g. ..._0001_000000.pgr or ...-0001-000000.pgr
PGR_FILENAME_RUN_PATTERN = re.compile(r"[-_](\d{4})[-_]0*000000\.pgr$", re.IGNORECASE)

# Generic fallback pattern used to pull a run number out of a folder name
# (raw GPS_Raw*.txt files live inside per-run folders, not per-run filenames).
GENERIC_RUN_NUMBER_PATTERN = re.compile(r"(\d{4})")

VIDEO_QUALITY_PRESETS = {
    "Low (640p, fast)":        {"scale": 0.5,  "bitrate_crf": 32, "fourcc": "mp4v"},
    "Medium (960p, balanced)": {"scale": 0.75, "bitrate_crf": 26, "fourcc": "mp4v"},
    "High (Full res)":         {"scale": 1.0,  "bitrate_crf": 20, "fourcc": "mp4v"},
    "Ultra (Full res, best)":  {"scale": 1.0,  "bitrate_crf": 14, "fourcc": "avc1"},
}

# ── New performance constants ──────────────────────────────────
MAX_CACHED_FRAMES = 60          # Expanded LRU cache for decoded frames to support prefetching
FRAME_DECODE_QUALITY = 85       # JPEG quality for caching
GPS_DECIMATION_FACTOR = 5000    # Reduce GPS points if > this count
EXPORT_BATCH_SIZE = 50          # Batch export size
MEMORY_WARNING_THRESHOLD = 0.85  # 85% memory usage triggers cleanup

# ─────────────────────────────────────────────────────────────
# Simple memory tracker (no psutil dependency)
# ─────────────────────────────────────────────────────────────
class MemoryTracker:
    """Simple memory usage tracker using Python's built-in modules"""
    
    @staticmethod
    def get_memory_usage():
        """Get current memory usage in MB (cross-platform approximation)"""
        try:
            import resource
            # Unix-like systems
            return resource.getrusage(resource.RUSAGE_SELF).ru_maxrss / 1024
        except (ImportError, AttributeError):
            try:
                import ctypes
                # Windows
                kernel32 = ctypes.windll.kernel32
                class PROCESS_MEMORY_COUNTERS(ctypes.Structure):
                    _fields_ = [
                        ("cb", ctypes.c_ulong),
                        ("PageFaultCount", ctypes.c_ulong),
                        ("PeakWorkingSetSize", ctypes.c_size_t),
                        ("WorkingSetSize", ctypes.c_size_t),
                        ("QuotaPeakPagedPoolUsage", ctypes.c_size_t),
                        ("QuotaPagedPoolUsage", ctypes.c_size_t),
                        ("QuotaPeakNonPagedPoolUsage", ctypes.c_size_t),
                        ("QuotaNonPagedPoolUsage", ctypes.c_size_t),
                        ("PagefileUsage", ctypes.c_size_t),
                        ("PeakPagefileUsage", ctypes.c_size_t),
                    ]
                
                counters = PROCESS_MEMORY_COUNTERS()
                counters.cb = ctypes.sizeof(counters)
                if kernel32.GetProcessMemoryInfo(kernel32.GetCurrentProcess(), ctypes.byref(counters), counters.cb):
                    return counters.WorkingSetSize / (1024 * 1024)
            except (ImportError, AttributeError):
                pass
            
            # Fallback: estimate from garbage collector
            gc.collect()
            return 0
    
    @staticmethod
    def get_available_memory():
        """Get estimated available memory in MB"""
        try:
            import psutil
            return psutil.virtual_memory().available / (1024 * 1024)
        except ImportError:
            # Fallback estimate
            return 1024  # Assume 1GB available

# ─────────────────────────────────────────────────────────────
# LRU Cache for decoded frames
# ─────────────────────────────────────────────────────────────
class LRUFrameCache:
    """Thread-safe LRU cache for decoded PIL images"""
    def __init__(self, max_size=MAX_CACHED_FRAMES):
        self.cache = OrderedDict()
        self.max_size = max_size
        self.lock = threading.Lock()
        self.hits = 0
        self.misses = 0
    
    def get(self, key):
        with self.lock:
            if key in self.cache:
                self.cache.move_to_end(key)
                self.hits += 1
                return self.cache[key]
            self.misses += 1
            return None
    
    def put(self, key, value):
        with self.lock:
            if key in self.cache:
                self.cache.move_to_end(key)
            self.cache[key] = value
            if len(self.cache) > self.max_size:
                oldest = next(iter(self.cache))
                del self.cache[oldest]
    
    def clear(self):
        with self.lock:
            self.cache.clear()
            self.hits = 0
            self.misses = 0
            gc.collect()
    
    def get_stats(self):
        with self.lock:
            total = self.hits + self.misses
            hit_rate = (self.hits / total * 100) if total > 0 else 0
            return {"size": len(self.cache), "hits": self.hits, "misses": self.misses, "hit_rate": hit_rate}
    
    def invalidate(self, key_pattern=None):
        with self.lock:
            if key_pattern is None:
                self.cache.clear()
            else:
                keys = list(self.cache.keys())
                for key in keys:
                    if key_pattern in str(key):
                        del self.cache[key]
            gc.collect()

# ─────────────────────────────────────────────────────────────
# Low-level PGR helpers (optimized for large files)
# ─────────────────────────────────────────────────────────────

def _swab32(n: int) -> int:
    return (
        ((n >> 24) & 0x000000FF) |
        ((n >>  8) & 0x0000FF00) |
        ((n <<  8) & 0x00FF0000) |
        ((n << 24) & 0xFF000000)
    )


def _decode_jpeg12(raw: bytes, quality=None) -> np.ndarray:
    """Optimized JPEG12 decoding with optional quality reduction"""
    if not HAVE_CODECS:
        raise RuntimeError("imagecodecs required for PGR decoding")
    arr = imagecodecs.jpeg_decode(raw)
    if arr.dtype == np.uint16:
        arr = np.right_shift(arr, 4).astype(np.uint8)
    else:
        arr = arr.astype(np.uint8)
    
    # Optional: reduce quality for large exports
    if quality and quality < 100 and arr.shape[0] > 1000:
        scale = quality / 100.0
        new_h, new_w = int(arr.shape[0] * scale), int(arr.shape[1] * scale)
        arr = np.array(Image.fromarray(arr).resize((new_w, new_h), Image.BILINEAR))
    
    return arr


def parse_frame_planes_fast(data: bytes, frame_start: int, total_size: int) -> list:
    """Optimized frame plane parsing with bounds checking"""
    if frame_start < 0 or frame_start + PGR_FRAME_HEADER_SIZE > total_size:
        return []
    sig_raw = struct.unpack_from("<I", data, frame_start + 16)[0]
    if _swab32(sig_raw) != PGR_CAFEBABE_VALUE:
        return []
    
    planes = [None] * PGR_MAX_SUBIMAGES
    table_base = frame_start + PGR_IMAGE_TABLE_OFFSET
    
    for i in range(PGR_MAX_SUBIMAGES):
        entry = table_base + i * 8
        if entry + 8 > total_size:
            break
        off_raw  = struct.unpack_from("<I", data, entry)[0]
        size_raw = struct.unpack_from("<I", data, entry + 4)[0]
        img_off  = _swab32(off_raw)
        img_size = _swab32(size_raw)
        
        if img_size < PGR_MIN_SUBIMAGE_BYTES:
            continue
            
        abs_start = frame_start + img_off
        abs_end   = abs_start + img_size
        
        if abs_end > total_size or data[abs_start:abs_start + 2] != b"\xff\xd8":
            continue
            
        planes[i] = (abs_start, abs_end, img_size)
    
    return planes


def scan_pgr_frames_optimized(filepath: Path, progress_cb=None) -> tuple:
    """
    Memory-map the PGR file with optimized scanning.
    Returns (mmap_obj, frame_offsets_list) - minimal metadata only.
    """
    file_size = filepath.stat().st_size
    f = open(filepath, "rb")
    try:
        mm = mmap.mmap(f.fileno(), 0, access=mmap.ACCESS_READ)
    except Exception:
        f.close()
        raise

    frames = []  # Store only (frame_start, plane_offsets)
    pos = 0
    total = len(mm)

    while True:
        idx = mm.find(CAFEBABE_NEEDLE, pos)
        if idx == -1:
            break
        
        frame_start = idx - 16
        header_end = frame_start + PGR_FRAME_HEADER_SIZE
        
        if header_end > total:
            pos = idx + 4
            continue
        
        # Quick signature validation
        sig_raw = struct.unpack_from("<I", mm, frame_start + 16)[0]
        if _swab32(sig_raw) != PGR_CAFEBABE_VALUE:
            pos = idx + 4
            continue
        
        # Parse plane offsets (store minimal info)
        planes = []
        valid_planes = 0
        table_base = frame_start + PGR_IMAGE_TABLE_OFFSET
        
        for i in range(PGR_MAX_SUBIMAGES):
            entry = table_base + i * 8
            if entry + 8 > total:
                break
            
            off_raw = struct.unpack_from("<I", mm, entry)[0]
            size_raw = struct.unpack_from("<I", mm, entry + 4)[0]
            img_off = _swab32(off_raw)
            img_size = _swab32(size_raw)
            
            if img_size < PGR_MIN_SUBIMAGE_BYTES:
                continue
                
            abs_start = frame_start + img_off
            abs_end = abs_start + img_size
            
            if abs_end > total:
                continue
                
            # Quick JPEG header check
            if mm[abs_start:abs_start + 2] != b"\xff\xd8":
                continue
                
            planes.append((abs_start, abs_end, img_size))  # Store size for quick access
            valid_planes += 1
        
        # Need at least camera 0's 4 planes to be valid
        if valid_planes >= PGR_PLANES_PER_CAMERA:
            frames.append({
                "frame_start": frame_start,
                "planes": planes,
                "file_offset": frame_start
            })
        
        pos = idx + 4
        
        if progress_cb and len(frames) % 50 == 0:
            pct = min((idx / total) * 100.0, 99.0)
            progress_cb(pct, f"Indexing… {len(frames)} frames")
    
    f.close()
    return mm, frames


# ─────────────────────────────────────────────────────────────
# Optimized frame decoding with caching
# ─────────────────────────────────────────────────────────────

def decode_camera_frame_optimized(
    mm: mmap.mmap,
    planes: list,
    camera_idx: int,
    r_gain: float = 1.0,
    g_gain: float = 1.0,
    b_gain: float = 1.0,
    rotate_270: bool = True,
    decode_quality: int = 100
) -> Image.Image:
    """Memory-efficient frame decoding with quality control"""
    base = camera_idx * PGR_PLANES_PER_CAMERA
    
    def get_plane_data(offset: int):
        for idx, (s, e, size) in enumerate(planes):
            if idx == base + offset:
                try:
                    return _decode_jpeg12(mm[s:e], decode_quality)
                except Exception:
                    return None
        return None
    
    R_plane = get_plane_data(0)
    G1_plane = get_plane_data(1)
    G2_plane = get_plane_data(2)
    B_plane = get_plane_data(3)
    
    ref = next((p for p in (R_plane, G1_plane, G2_plane, B_plane) if p is not None), None)
    if ref is None:
        raise ValueError(f"No valid planes for camera {camera_idx}")
    
    h, w = ref.shape[:2]
    zeros = np.zeros((h, w), dtype=np.uint8)
    
    R = R_plane if R_plane is not None else zeros
    G1 = G1_plane if G1_plane is not None else zeros
    G2 = G2_plane if G2_plane is not None else zeros
    B = B_plane if B_plane is not None else zeros
    
    G = ((G1.astype(np.int32) + G2.astype(np.int32)) >> 1).astype(np.uint8)
    
    if r_gain != 1.0:
        R = np.clip(R.astype(np.float32) * r_gain, 0, 255).astype(np.uint8)
    if g_gain != 1.0:
        G = np.clip(G.astype(np.float32) * g_gain, 0, 255).astype(np.uint8)
    if b_gain != 1.0:
        B = np.clip(B.astype(np.float32) * b_gain, 0, 255).astype(np.uint8)
    
    img = Image.fromarray(np.stack([R, G, B], axis=2), mode="RGB")
    if rotate_270:
        img = img.rotate(270, expand=True)
    return img


def decode_dual_panoramic_optimized(mm, planes: list, r_gain=1.0, g_gain=1.0, b_gain=1.0,
                                     overlap_percent=0.25, decode_quality=100) -> Image.Image:
    """Optimized dual panoramic decoding"""
    img_left = decode_camera_frame_optimized(mm, planes, 2, r_gain, g_gain, b_gain, True, decode_quality)
    img_right = decode_camera_frame_optimized(mm, planes, 3, r_gain, g_gain, b_gain, True, decode_quality)
    
    # Convert to numpy efficiently
    arr_l = np.asarray(img_left, dtype=np.float32)
    arr_r = np.asarray(img_right, dtype=np.float32)
    H_l, W_l = arr_l.shape[:2]
    H_r, W_r = arr_r.shape[:2]
    
    if H_l != H_r:
        target_h = max(H_l, H_r)
        if H_l < target_h:
            img_left = img_left.resize((W_l, target_h), Image.BILINEAR)
            arr_l = np.asarray(img_left, dtype=np.float32)
        else:
            img_right = img_right.resize((W_r, target_h), Image.BILINEAR)
            arr_r = np.asarray(img_right, dtype=np.float32)
    
    V_SHIFT = 12
    if V_SHIFT > 0:
        arr_l = arr_l[:H_l - V_SHIFT, :]
        arr_r = arr_r[V_SHIFT:, :]
    
    work_h = min(arr_l.shape[0], arr_r.shape[0])
    arr_l = arr_l[:work_h, :]
    arr_r = arr_r[:work_h, :]
    W_l = arr_l.shape[1]
    W_r = arr_r.shape[1]
    
    overlap_w = max(1, int(W_l * overlap_percent))
    out_w = W_l + W_r - overlap_w
    output = np.empty((work_h, out_w, 3), dtype=np.float32)
    output[:, :W_l - overlap_w, :] = arr_l[:, :W_l - overlap_w, :]
    output[:, W_l:, :] = arr_r[:, overlap_w:, :]
    
    # Feather blend
    left_strip = arr_l[:, W_l - overlap_w:W_l, :]
    right_strip = arr_r[:, :overlap_w, :]
    feather = min(20, overlap_w)
    alpha_r = np.ones(overlap_w, dtype=np.float32)
    alpha_r[:feather] = np.linspace(0.0, 1.0, feather)
    alpha_l = 1.0 - alpha_r
    output[:, W_l - overlap_w:W_l, :] = (
        left_strip * alpha_l[np.newaxis, :, np.newaxis] +
        right_strip * alpha_r[np.newaxis, :, np.newaxis]
    )
    
    return Image.fromarray(np.clip(output, 0, 255).astype(np.uint8), mode="RGB")


def decode_all_cameras_grid_optimized(mm, planes: list,
                                       r_gain=1.0, g_gain=1.0, b_gain=1.0,
                                       decode_quality=100) -> Image.Image:
    """Optimized 6-camera grid decoding"""
    images = [decode_camera_frame_optimized(mm, planes, c, r_gain, g_gain, b_gain, True, decode_quality)
              for c in range(PGR_NUM_CAMERAS)]
    w, h = images[0].size
    grid_img = Image.new("RGB", (w * 3, h * 2))
    for idx, img in enumerate(images):
        grid_img.paste(img, ((idx % 3) * w, (idx // 3) * h))
    return grid_img


# ─────────────────────────────────────────────────────────────
# Optimized GPS handling for large datasets
# ─────────────────────────────────────────────────────────────

def _haversine_m(lat1, lon1, lat2, lon2) -> float:
    R = 6_371_000.0
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dp = math.radians(lat2 - lat1)
    dl = math.radians(lon2 - lon1)
    a = math.sin(dp/2)**2 + math.cos(phi1) * math.cos(phi2) * math.sin(dl/2)**2
    return 2 * R * math.atan2(math.sqrt(a), math.sqrt(1 - a))


def decimate_gps_points(coords: list, max_points: int = 5000) -> list:
    """Reduce GPS points to manageable size using adaptive sampling"""
    if len(coords) <= max_points:
        return coords
    
    step = len(coords) / max_points
    result = []
    for i in range(max_points):
        idx = int(i * step)
        if idx < len(coords):
            result.append(coords[idx])
    
    # Always include last point
    if result[-1] != coords[-1]:
        result.append(coords[-1])
    
    return result


def parse_coord_text_optimized(text: str) -> list:
    """Optimized coordinate parser"""
    result = []
    for pt in text.strip().split():
        parts = pt.strip().split(",")
        if len(parts) < 2:
            continue
        try:
            lon = float(parts[0])
            lat = float(parts[1])
            alt = float(parts[2]) if len(parts) >= 3 else 0.0
            if lat != 0.0 or lon != 0.0:
                result.append({"lat": lat, "lon": lon, "altitude_m": alt, "name": "", "timestamp": None})
        except ValueError:
            pass
    return result


def extract_gps_by_run_from_kml(kml_path: str) -> dict:
    """Parse KML/KMZ and return GPS points grouped by run number.

    Returns a dict: {run_number_int: [coord_dict, ...], ...}
    The overview 'Trajectory_Path - 0001' placemark is always skipped.
    Placemarks not matching the 'Run NNNN' naming pattern are grouped
    under run number None (so nothing is silently dropped).
    """
    runs = {}
    try:
        if kml_path.lower().endswith(".kmz"):
            with zipfile.ZipFile(kml_path, "r") as kmz:
                kml_files = [n for n in kmz.namelist() if n.endswith(".kml")]
                if not kml_files:
                    return runs
                content = kmz.read(kml_files[0])
        else:
            with open(kml_path, "rb") as f:
                content = f.read()

        # Parse the full tree so we can walk Placemark -> name -> coordinates
        root = ET.fromstring(content)

        # Detect namespace prefix (handles both namespaced and bare KML)
        ns = ""
        if root.tag.startswith("{"):
            ns = root.tag.split("}")[0] + "}"

        def tag(local):
            return f"{ns}{local}"

        # Walk every Placemark in the document
        for placemark in root.iter(tag("Placemark")):
            # Read the Placemark name (try both namespaced and bare)
            name_elem = placemark.find(tag("name"))
            if name_elem is None:
                name_elem = placemark.find("name")
            raw_name = (name_elem.text or "").strip() if name_elem is not None else ""
            placemark_name = raw_name.lower()

            # Skip trajectory_path-0001 entirely
            if placemark_name in STRIP_PLACEMARK_NAMES:
                continue

            # Determine which run this placemark belongs to
            run_match = RUN_PLACEMARK_PATTERN.search(raw_name)
            run_number = int(run_match.group(1)) if run_match else None

            # Extract all <coordinates> blocks inside this Placemark
            pm_coords = []
            for coord_elem in placemark.iter(tag("coordinates")):
                if coord_elem.text:
                    pm_coords.extend(parse_coord_text_optimized(coord_elem.text))

            if not pm_coords:
                continue

            runs.setdefault(run_number, []).extend(pm_coords)

    except Exception as e:
        print(f"Error parsing KML: {e}")

    return runs


def extract_gps_from_kml_optimized(kml_path: str) -> list:
    """Flat-list GPS extraction (back-compat helper). Used where a single
    combined track is needed (e.g. CSV/KML export of everything loaded)."""
    runs = extract_gps_by_run_from_kml(kml_path)
    flat = []
    for run_number in sorted(runs.keys(), key=lambda x: (x is None, x)):
        flat.extend(runs[run_number])
    return flat


def extract_multiple_gps_optimized(kml_paths: list) -> list:
    """Process multiple GPS files with progress reporting"""
    all_coords = []
    total_files = len(kml_paths)
    
    for i, path in enumerate(kml_paths):
        coords = extract_gps_from_kml_optimized(path)
        all_coords.extend(coords)
        
        # Progress reporting
        if total_files > 1 and i % 10 == 0:
            print(f"GPS Load Progress: {i+1}/{total_files} files, {len(all_coords)} points")
    
    return all_coords


def extract_multiple_gps_by_run(kml_paths: list) -> dict:
    """Process multiple GPS files, merging their per-run dicts together.
    If the same run number appears in more than one file, points are appended
    in file order (this matches how multiple KMZ coverage files would be loaded).
    """
    merged_runs = {}
    for path in kml_paths:
        runs = extract_gps_by_run_from_kml(path)
        for run_number, pts in runs.items():
            merged_runs.setdefault(run_number, []).extend(pts)
    return merged_runs


def extract_run_number_from_pgr_filename(filename: str) -> "int | None":
    """Extract the 4-digit run number embedded in a PGR filename, e.g.
    'TMX60326010801-000125-0001-000000.pgr' -> 1
    Returns None if no run number pattern is found.
    """
    match = PGR_FILENAME_RUN_PATTERN.search(filename)
    if match:
        return int(match.group(1))
    return None


def extract_run_number_from_raw_gps_path(folder_name: str, file_name: str) -> "int | None":
    """Try to find a run number for a GPS_Raw*.txt file based on the name of
    its parent folder first (typical layout: .../Run 0001/GPS_Raw....txt),
    falling back to the filename itself, and finally to any 4-digit group
    found anywhere in the folder/file name."""
    for candidate in (folder_name, file_name):
        if not candidate:
            continue
        m = RUN_PLACEMARK_PATTERN.search(candidate)
        if m:
            return int(m.group(1))
        m = PGR_FILENAME_RUN_PATTERN.search(candidate)
        if m:
            return int(m.group(1))
    for candidate in (folder_name, file_name):
        if not candidate:
            continue
        m = GENERIC_RUN_NUMBER_PATTERN.search(candidate)
        if m:
            return int(m.group(1))
    return None


def nmea_to_decimal(coord, direction):
    """Convert NMEA coordinate (DDMM.MMMM) to decimal degrees.
    (Ported verbatim from CombineTextFiles1.py.)"""
    if not coord:
        return None
    try:
        coord = float(coord)
    except (TypeError, ValueError):
        return None
    degrees = int(coord / 100)
    minutes = coord - (degrees * 100)
    decimal = degrees + minutes / 60
    if direction in ("S", "W"):
        decimal *= -1
    return decimal


def parse_gps_raw_txt_file(file_path: Path) -> list:
    """Parse a single GPS_Raw*.txt file (one JSON record per line, each
    holding an 'OdoDataRecord' block and a raw 'NmeaLine') into a flat list
    of coord dicts compatible with the rest of the GPS pipeline:
    {"lat", "lon", "altitude_m", "name", "timestamp", "chainage", "speed"}.

    This is the same parsing logic as CombineTextFiles1.py, just returning
    structured records in-memory instead of writing straight to CSV.
    """
    records = []
    try:
        with open(file_path, "r", encoding="utf-8", errors="ignore") as infile:
            for line in infile:
                try:
                    record = json.loads(line.strip())
                    odo = record.get("OdoDataRecord", {})
                    nmea = record.get("NmeaLine", "")

                    parts = nmea.split(",")
                    if not parts:
                        continue
                    if parts[0].endswith("GGA") or parts[0].endswith("RMC"):
                        lat = nmea_to_decimal(parts[2], parts[3]) if len(parts) > 3 else None
                        lon = nmea_to_decimal(parts[4], parts[5]) if len(parts) > 5 else None
                        if lat is None or lon is None:
                            continue
                        if lat == 0.0 and lon == 0.0:
                            continue

                        records.append({
                            "lat": lat,
                            "lon": lon,
                            "altitude_m": 0.0,
                            "name": "",
                            "timestamp": odo.get("Time"),
                            "chainage": odo.get("Chainage", ""),
                            "speed": odo.get("Speed", ""),
                        })
                except Exception:
                    continue
    except Exception as e:
        print(f"Error parsing GPS_Raw file {file_path}: {e}")
    return records


def extract_gps_by_run_from_raw_txt_folder(root_folder: str, progress_cb=None) -> dict:
    """Walk root_folder recursively for GPS_Raw*.txt files (same discovery
    rule as CombineTextFiles1.py: filename startswith 'GPS_Raw' and ends
    with '.txt'), parse each one, and group the resulting coord dicts by
    run number (matching the same grouping convention used for KML/KMZ:
    {run_number_int_or_None: [coord_dict, ...]})."""
    runs = {}
    txt_files = []
    for foldername, _subfolders, filenames in os.walk(root_folder):
        for filename in filenames:
            if filename.startswith("GPS_Raw") and filename.endswith(".txt"):
                txt_files.append((foldername, filename))

    total = len(txt_files)
    for i, (foldername, filename) in enumerate(txt_files):
        file_path = os.path.join(foldername, filename)
        parent_folder = os.path.basename(foldername)
        run_number = extract_run_number_from_raw_gps_path(parent_folder, filename)

        pts = parse_gps_raw_txt_file(Path(file_path))
        if pts:
            runs.setdefault(run_number, []).extend(pts)

        if progress_cb and total:
            progress_cb(i + 1, total, sum(len(v) for v in runs.values()))

    return runs


def export_raw_gps_runs_to_csv(runs: dict, output_csv: str):
    """Write parsed GPS_Raw data to a CSV in the exact column layout used by
    CombineTextFiles1.py (Parent_Folder, File, Time, Chainage, Speed,
    Latitude, Longitude). 'Parent_Folder'/'File' are reconstructed from the
    run grouping since by that point individual file provenance per-point
    isn't tracked separately; Run number is used in their place."""
    with open(output_csv, "w", newline="", encoding="utf-8") as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(["Run", "Time", "Chainage", "Speed", "Latitude", "Longitude"])
        for run_number in sorted(runs.keys(), key=lambda x: (x is None, x)):
            run_label = f"{run_number:04d}" if run_number is not None else ""
            for c in runs[run_number]:
                writer.writerow([
                    run_label,
                    c.get("timestamp", ""),
                    c.get("chainage", ""),
                    c.get("speed", ""),
                    c.get("lat", ""),
                    c.get("lon", ""),
                ])


def _catmull_rom_segment(p0, p1, p2, p3, steps=8):
    """Catmull-Rom spline interpolation between p1 and p2 using p0 and p3 as tangent guides.
    Returns a list of (lat, lon) tuples including p1 but not p2.
    """
    pts = []
    for i in range(steps):
        t = i / steps
        t2 = t * t
        t3 = t2 * t
        # Catmull-Rom basis coefficients
        h00 = 2*t3 - 3*t2 + 1
        h10 = t3 - 2*t2 + t
        h01 = -2*t3 + 3*t2
        h11 = t3 - t2
        # Tangents (alpha=0.5 for centripetal Catmull-Rom feel)
        m1_lat = 0.5 * (p2[0] - p0[0])
        m1_lon = 0.5 * (p2[1] - p0[1])
        m2_lat = 0.5 * (p3[0] - p1[0])
        m2_lon = 0.5 * (p3[1] - p1[1])
        lat = h00*p1[0] + h10*m1_lat + h01*p2[0] + h11*m2_lat
        lon = h00*p1[1] + h10*m1_lon + h01*p2[1] + h11*m2_lon
        pts.append((lat, lon))
    return pts


def _smooth_segment(seg, steps=8):
    """Apply Catmull-Rom smoothing to a segment of (lat, lon) points.
    Segments with fewer than 2 points are returned unchanged.
    """
    n = len(seg)
    if n < 2:
        return seg
    if n == 2:
        return seg
    smoothed = []
    for i in range(n - 1):
        p0 = seg[max(0, i - 1)]
        p1 = seg[i]
        p2 = seg[i + 1]
        p3 = seg[min(n - 1, i + 2)]
        smoothed.extend(_catmull_rom_segment(p0, p1, p2, p3, steps=steps))
    smoothed.append(seg[-1])  # Always include the final point
    return smoothed


def build_gps_segments_optimized(coords: list) -> list:
    """Build GPS path segments with max gap = GPS_MAX_SEGMENT_M (50 m) and
    Catmull-Rom smoothing applied so the rendered track looks natural.
    """
    if not coords:
        return []

    segments = []
    current = [coords[0]]

    for i in range(1, len(coords)):
        dist = _haversine_m(coords[i-1][0], coords[i-1][1], coords[i][0], coords[i][1])
        if dist <= GPS_MAX_SEGMENT_M:
            current.append(coords[i])
        else:
            if len(current) >= 2:
                segments.append(_smooth_segment(current))
            current = [coords[i]]

    if len(current) >= 2:
        segments.append(_smooth_segment(current))

    return segments


# ─────────────────────────────────────────────────────────────
# Streaming Excel Export (memory efficient)
# ─────────────────────────────────────────────────────────────

class StreamingExcelExporter:
    """Write Excel files incrementally to avoid memory issues"""
    
    def __init__(self, output_path: Path):
        self.output_path = output_path
        self.buffer = []
        self.batch_size = EXPORT_BATCH_SIZE
        self.writer = None
        self.workbook = None
        self.sheet = None
    
    def start(self):
        import openpyxl
        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = "Survey_Data"
        self._write_header()
    
    def _write_header(self):
        headers = ["File", "Frame", "Latitude", "Longitude", "Altitude_m", "Point_Name", "Chainage_m"]
        for col, header in enumerate(headers, 1):
            self.sheet.cell(row=1, column=col, value=header)
        self.current_row = 2
    
    def add_row(self, row_data: dict):
        """Add a row to buffer, flush when full"""
        self.buffer.append(row_data)
        if len(self.buffer) >= self.batch_size:
            self._flush_buffer()
    
    def _flush_buffer(self):
        """Write buffered rows to Excel"""
        for row_data in self.buffer:
            self.sheet.cell(row=self.current_row, column=1, value=row_data.get("File", ""))
            self.sheet.cell(row=self.current_row, column=2, value=row_data.get("Frame", ""))
            self.sheet.cell(row=self.current_row, column=3, value=row_data.get("Latitude", ""))
            self.sheet.cell(row=self.current_row, column=4, value=row_data.get("Longitude", ""))
            self.sheet.cell(row=self.current_row, column=5, value=row_data.get("Altitude_m", ""))
            self.sheet.cell(row=self.current_row, column=6, value=row_data.get("Point_Name", ""))
            self.sheet.cell(row=self.current_row, column=7, value=row_data.get("Chainage_m", ""))
            self.current_row += 1
        self.buffer.clear()
    
    def finish(self):
        """Final flush and save"""
        if self.buffer:
            self._flush_buffer()
        
        # Auto-adjust column widths
        for col in self.sheet.columns:
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            self.sheet.column_dimensions[column_letter].width = adjusted_width
        
        self.workbook.save(self.output_path)
        self.workbook.close()


# ─────────────────────────────────────────────────────────────
# CheckboxDropdown - Custom dropdown with checkboxes
# ─────────────────────────────────────────────────────────────

class CheckboxDropdown:
    """Custom dropdown widget with checkboxes for multi-selection"""
    
    def __init__(self, parent, values, on_select_callback=None):
        self.parent = parent
        self.values = values
        self.selected = {val: True for val in values} if values else {}
        self.on_select_callback = on_select_callback
        
        # Main frame
        self.frame = tk.Frame(parent, bg=BG_CARD, relief="flat", bd=1, highlightthickness=1, highlightbackground="#2a3040")
        
        # Button to show dropdown
        self.button = tk.Button(
            self.frame, 
            text="Select Files...", 
            bg=BG_PANEL, 
            fg=ACCENT,
            font=("Segoe UI", 9),
            relief="flat",
            cursor="hand2",
            command=self.toggle_dropdown
        )
        self.button.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=4, pady=2)
        
        # Label to show selection summary
        self.summary_label = tk.Label(
            self.frame, 
            text="All files", 
            bg=BG_CARD, 
            fg=FG_DIM, 
            font=("Segoe UI", 8),
            anchor="w"
        )
        self.summary_label.pack(side=tk.RIGHT, padx=4)
        
        # Dropdown window (initially hidden)
        self.dropdown = None
        self.check_vars = {}
        
        self.update_summary()
    
    def set_values(self, values):
        """Update the list of values"""
        self.values = values
        self.selected = {val: True for val in values} if values else {}
        self.check_vars.clear()
        self.update_summary()
        self._refresh_dropdown()
    
    def update_summary(self):
        """Update the summary label text"""
        selected_count = sum(1 for v in self.selected.values() if v)
        total_count = len(self.values)
        
        if selected_count == 0:
            self.summary_label.config(text="No files selected", fg=DANGER)
        elif selected_count == total_count:
            self.summary_label.config(text=f"All {total_count} files", fg=SUCCESS)
        else:
            self.summary_label.config(text=f"{selected_count}/{total_count} files", fg=ACCENT)
    
    def toggle_dropdown(self):
        """Show or hide the dropdown menu"""
        if self.dropdown and self.dropdown.winfo_exists():
            self.hide_dropdown()
        else:
            self.show_dropdown()
    
    def show_dropdown(self):
        """Create and show the dropdown menu"""
        if self.dropdown and self.dropdown.winfo_exists():
            self.dropdown.destroy()
        
        # Get button position
        x = self.button.winfo_rootx()
        y = self.button.winfo_rooty() + self.button.winfo_height()
        
        self.dropdown = tk.Toplevel(self.parent)
        self.dropdown.overrideredirect(True)
        self.dropdown.configure(bg=BG_CARD, bd=1, highlightthickness=1, highlightbackground="#2a3040")
        self.dropdown.geometry(f"+{x}+{y}")
        
        # Create scrollable frame
        canvas = tk.Canvas(self.dropdown, bg=BG_CARD, highlightthickness=0)
        scrollbar = tk.Scrollbar(self.dropdown, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg=BG_CARD)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Header with Select All / Clear All buttons
        header_frame = tk.Frame(scrollable_frame, bg=BG_CARD)
        header_frame.pack(fill=tk.X, padx=6, pady=4)
        
        select_all_btn = tk.Button(
            header_frame, 
            text="✓ Select All", 
            bg="#0a4d36", 
            fg=SUCCESS,
            font=("Segoe UI", 8, "bold"),
            relief="flat",
            cursor="hand2",
            command=self.select_all
        )
        select_all_btn.pack(side=tk.LEFT, padx=2)
        
        clear_all_btn = tk.Button(
            header_frame, 
            text="✗ Clear All", 
            bg="#6b3200", 
            fg=WARNING,
            font=("Segoe UI", 8, "bold"),
            relief="flat",
            cursor="hand2",
            command=self.clear_all
        )
        clear_all_btn.pack(side=tk.LEFT, padx=2)
        
        # Separator
        tk.Frame(header_frame, bg="#2a3040", height=1).pack(fill=tk.X, pady=4)
        
        # Add checkboxes for each file
        for value in self.values:
            var = tk.BooleanVar(value=self.selected.get(value, True))
            self.check_vars[value] = var
            
            cb_frame = tk.Frame(scrollable_frame, bg=BG_CARD)
            cb_frame.pack(fill=tk.X, padx=8, pady=2)
            
            cb = tk.Checkbutton(
                cb_frame,
                text=value,
                variable=var,
                bg=BG_CARD,
                fg=FG_MAIN,
                selectcolor=BG_CARD,
                activebackground=BG_CARD,
                font=("Segoe UI", 9),
                cursor="hand2",
                command=lambda v=value, vvar=var: self.on_check_changed(v, vvar)
            )
            cb.pack(side=tk.LEFT, anchor="w")
        
        # Apply button
        apply_frame = tk.Frame(scrollable_frame, bg=BG_CARD)
        apply_frame.pack(fill=tk.X, padx=6, pady=4)
        
        tk.Frame(apply_frame, bg="#2a3040", height=1).pack(fill=tk.X, pady=4)
        
        apply_btn = tk.Button(
            apply_frame,
            text="✓ Apply Selection",
            bg=BTN_PRIMARY,
            fg="white",
            font=("Segoe UI", 9, "bold"),
            relief="flat",
            cursor="hand2",
            command=self.apply_selection
        )
        apply_btn.pack(fill=tk.X, pady=2)
        
        cancel_btn = tk.Button(
            apply_frame,
            text="Cancel",
            bg=BG_PANEL,
            fg=FG_DIM,
            font=("Segoe UI", 9),
            relief="flat",
            cursor="hand2",
            command=self.hide_dropdown
        )
        cancel_btn.pack(fill=tk.X, pady=2)
        
        # Layout
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Set max height
        self.dropdown.update_idletasks()
        max_height = min(400, scrollable_frame.winfo_reqheight() + 100)
        self.dropdown.geometry(f"300x{max_height}+{x}+{y}")
        
        # Close dropdown when clicking outside
        self.dropdown.focus_set()
        self.dropdown.bind("<FocusOut>", lambda e: self.hide_dropdown())
    
    def hide_dropdown(self):
        """Destroy the dropdown window"""
        if self.dropdown and self.dropdown.winfo_exists():
            self.dropdown.destroy()
        self.dropdown = None
    
    def on_check_changed(self, value, var):
        """Handle individual checkbox changes"""
        self.selected[value] = var.get()
    
    def select_all(self):
        """Select all checkboxes"""
        for value, var in self.check_vars.items():
            var.set(True)
            self.selected[value] = True
        self.update_summary()
        self._refresh_dropdown()
    
    def clear_all(self):
        """Clear all checkboxes"""
        for value, var in self.check_vars.items():
            var.set(False)
            self.selected[value] = False
        self.update_summary()
        self._refresh_dropdown()
    
    def apply_selection(self):
        """Apply the current selection and close dropdown"""
        # Update selected from current vars
        for value, var in self.check_vars.items():
            self.selected[value] = var.get()
        self.update_summary()
        self.hide_dropdown()
        
        # Call callback with selected items
        if self.on_select_callback:
            selected_items = [v for v, s in self.selected.items() if s]
            self.on_select_callback(selected_items)
    
    def _refresh_dropdown(self):
        """Refresh the dropdown if it's open"""
        if self.dropdown and self.dropdown.winfo_exists():
            self.hide_dropdown()
            self.show_dropdown()
    
    def get_selected(self):
        """Get list of selected values"""
        return [v for v, s in self.selected.items() if s]
    
    def pack(self, **kwargs):
        self.frame.pack(**kwargs)
    
    def pack_forget(self):
        self.frame.pack_forget()


# ─────────────────────────────────────────────────────────────
# MergedPGRIndex with lazy loading
# ─────────────────────────────────────────────────────────────

class MergedPGRIndex:
    """Manages merged PGR files with lazy frame access"""
    
    def __init__(self):
        self.file_entries = []
        self.virtual_table = []
        self._frame_cache = LRUFrameCache()
    
    def clear(self):
        self.file_entries = []
        self.virtual_table = []
        self._frame_cache.clear()
        gc.collect()
    
    def add_file(self, path: Path, mm, frames: list):
        fe_idx = len(self.file_entries)
        self.file_entries.append({
            "path": path,
            "mm": mm,
            "frames": frames,
            "mmap_ref": weakref.ref(mm)
        })
        for local_idx in range(len(frames)):
            self.virtual_table.append((fe_idx, local_idx))
    
    def add_selected_files(self, file_paths: list, file_cache: dict):
        """Add only selected files to the merge index"""
        added_count = 0
        for fp in file_paths:
            cache_entry = file_cache.get(fp)
            if cache_entry and cache_entry["frames"]:
                self.add_file(fp, cache_entry["mm"], cache_entry["frames"])
                added_count += 1
        return added_count
    
    @property
    def total_frames(self):
        return len(self.virtual_table)
    
    def get_frame_data(self, virtual_idx: int):
        fe_idx, local_idx = self.virtual_table[virtual_idx]
        entry = self.file_entries[fe_idx]
        return entry["mm"], entry["frames"][local_idx]
    
    def file_name_at(self, virtual_idx: int) -> str:
        fe_idx, _ = self.virtual_table[virtual_idx]
        return self.file_entries[fe_idx]["path"].name
    
    def get_cache_key(self, virtual_idx: int, view_mode: str, camera: int, gains: tuple) -> str:
        """Generate cache key for frame"""
        return f"{virtual_idx}_{view_mode}_{camera}_{gains}"
    
    def cache_frame(self, key: str, image: Image.Image):
        self._frame_cache.put(key, image)
    
    def get_cached_frame(self, key: str):
        return self._frame_cache.get(key)
    
    def get_cache_stats(self):
        return self._frame_cache.get_stats()
    
    def get_file_names(self):
        """Get list of file names in merge order"""
        return [entry["path"].name for entry in self.file_entries]


# ─────────────────────────────────────────────────────────────
# Main Application
# ─────────────────────────────────────────────────────────────

# ─────────────────────────────────────────────────────────────
# DataModeDialog - startup selector: LiDAR (KMZ/KML GPS) vs LCMS (TXT GPS)
# ─────────────────────────────────────────────────────────────

class DataModeDialog:
    """Modal startup dialog that asks the user which survey data source
    they're working with before the main viewer opens:

      • LiDAR  -> GPS is supplied as KML/KMZ trajectory files
      • LCMS   -> GPS is supplied as GPS_Raw*.txt files (NMEA/Odo JSON lines)

    The choice is stored on the returned PGRViewerPro instance as
    `self.data_mode` and drives which GPS-loading button/workflow the
    toolbar exposes. Use DataModeDialog.ask(root) to show it and block
    until the user picks a mode (or closes/cancels, which returns None).
    """

    MODES = {
        "LiDAR  —  GPS from KML / KMZ": "LIDAR",
        "LCMS  —  GPS from TXT files": "LCMS",
    }

    def __init__(self, parent: tk.Tk):
        self.result = None

        self.win = tk.Toplevel(parent)
        self.win.title("PGR Survey Viewer — Select Data Source")
        self.win.configure(bg=BG_DARK)
        self.win.resizable(False, False)
        self.win.protocol("WM_DELETE_WINDOW", self._on_cancel)
        self.win.grab_set()

        # ── Header ──
        header = tk.Frame(self.win, bg=BG_DARK)
        header.pack(fill=tk.X, padx=24, pady=(22, 10))
        tk.Label(header, text="⬡", bg=BG_DARK, fg=ACCENT, font=("Segoe UI", 24)).pack(side=tk.LEFT, padx=(0, 10))
        title_box = tk.Frame(header, bg=BG_DARK)
        title_box.pack(side=tk.LEFT)
        tk.Label(title_box, text="PGR SURVEY VIEWER", bg=BG_DARK, fg=FG_BRIGHT, font=("Segoe UI", 15, "bold")).pack(anchor="w")
        tk.Label(title_box, text="Choose the survey data source to continue", bg=BG_DARK, fg=FG_DIM, font=("Segoe UI", 9)).pack(anchor="w")

        tk.Frame(self.win, bg=ACCENT, height=2).pack(fill=tk.X, padx=24)

        # ── Body ──
        body = tk.Frame(self.win, bg=BG_PANEL)
        body.pack(fill=tk.BOTH, expand=True, padx=24, pady=18)

        tk.Label(body, text="Data Type:", bg=BG_PANEL, fg=FG_MAIN, font=("Segoe UI", 10, "bold")).pack(anchor="w", pady=(0, 6))

        self.mode_var = tk.StringVar(value=list(self.MODES.keys())[0])
        combo = ttk.Combobox(body, textvariable=self.mode_var, values=list(self.MODES.keys()),
                              state="readonly", width=34, font=("Segoe UI", 10))
        combo.pack(anchor="w", fill=tk.X)
        combo.bind("<<ComboboxSelected>>", self._on_mode_selected)

        self.desc_label = tk.Label(body, text="", bg=BG_PANEL, fg=FG_DIM, font=("Segoe UI", 8), justify="left", wraplength=380)
        self.desc_label.pack(anchor="w", pady=(8, 0))
        self._on_mode_selected()

        # ── Footer buttons ──
        footer = tk.Frame(self.win, bg=BG_DARK)
        footer.pack(fill=tk.X, padx=24, pady=(4, 22))

        tk.Button(footer, text="Exit", bg=BG_CARD, fg=FG_DIM, font=("Segoe UI", 9), relief="flat",
                  bd=0, padx=14, pady=8, cursor="hand2", command=self._on_cancel).pack(side=tk.RIGHT)
        ttk.Button(footer, text="Open PGR Viewer  ▶", style="Primary.TButton",
                   command=self._on_confirm).pack(side=tk.RIGHT, padx=(0, 8))

        self.win.update_idletasks()
        w, h = 460, 260
        sw, sh = self.win.winfo_screenwidth(), self.win.winfo_screenheight()
        self.win.geometry(f"{w}x{h}+{(sw - w) // 2}+{(sh - h) // 2}")

    def _on_mode_selected(self, event=None):
        code = self.MODES.get(self.mode_var.get(), "LIDAR")
        if code == "LIDAR":
            self.desc_label.config(
                text="LiDAR mode: load GPS trajectories from KML/KMZ coverage files "
                     "(Load GPS (LiDAR KML/KMZ) button).")
        else:
            self.desc_label.config(
                text="LCMS mode: load GPS from GPS_Raw*.txt export files "
                     "(Load GPS (LCMS TXT Folder) button).")

    def _on_confirm(self):
        self.result = self.MODES.get(self.mode_var.get(), "LIDAR")
        self.win.grab_release()
        self.win.destroy()

    def _on_cancel(self):
        self.result = None
        self.win.grab_release()
        self.win.destroy()

    @staticmethod
    def ask(parent: tk.Tk):
        """Show the dialog modally and return 'LIDAR', 'LCMS', or None if cancelled."""
        dlg = DataModeDialog(parent)
        parent.wait_window(dlg.win)
        return dlg.result


class PGRViewerPro:
    def __init__(self, root: tk.Tk, data_mode: str = "LIDAR"):
        self.root = root
        self.data_mode = data_mode if data_mode in ("LIDAR", "LCMS") else "LIDAR"
        self.root.title(f"PGR Survey Viewer v7  ·  {self.data_mode} Mode")
        self.root.geometry("1680x960")
        self.root.minsize(1280, 720)
        self.root.configure(bg=BG_DARK)
        
        # ── Thread Pools for Background Operations ─────────────
        self.prefetch_executor = ThreadPoolExecutor(max_workers=max(2, (os.cpu_count() or 4) - 1))
        
        # ── PGR state (optimized) ──────────────────────────────
        self.pgr_files: list = []
        self.file_cache: dict = {}
        self.merged_index = MergedPGRIndex()
        self.current_file_idx = 0
        self.current_frame_idx = 0
        self.total_frames = 0
        self.is_playing = False
        self._after_id = None
        self.is_indexing = False
        self.merged_mode = False
        self._single_mm = None
        self._single_frames = []
        self.selected_merge_files = []  # List of selected files for merge
        
        # ── Display ────────────────────────────────────────────
        self.current_cam_idx = 2
        self.view_mode = "single"
        self.r_gain = 1.1
        self.g_gain = 1.0
        self.b_gain = 1.0
        self.decode_quality = 100  # Dynamic quality adjustment
        self.last_frame_time = 0
        self._current_pil_image = None
        
        # ── Zoom / Pan state (GR view panel) ────────────────────
        self.zoom_level = 1.0       # 1.0 = fit-to-window ("same view")
        self.zoom_min = 1.0
        self.zoom_max = 12.0
        self.zoom_step = 1.18
        self.pan_offset_x = 0.0     # canvas-space panning offset (pixels)
        self.pan_offset_y = 0.0
        self._is_panning = False
        self._pan_start = None
        self._resize_after_id = None
        
        # ── GPS state (optimized) ──────────────────────────────
        self.gps_coords = []           # Flat (lat, lon) list - ALL loaded runs (legacy/export use)
        self.gps_full_data = []        # Flat full coord dicts - ALL loaded runs (legacy/export use)
        self.gps_frame_map = {}        # frame_idx -> (lat, lon) for the CURRENTLY ACTIVE selection only
        self.map_paths = []
        self.car_marker = None
        self.current_gps_files = []
        self.gps_by_run = {}           # run_number(int|None) -> [coord_dict, ...] (raw parsed data, all runs)
        self.active_gps_coords = []    # (lat, lon) list for whatever is CURRENTLY shown/mapped
        self.active_gps_full_data = [] # full coord dicts matching active_gps_coords
        self.active_run_label = ""     # human-readable label of what's currently active ("All runs" / "Run 0003" / etc.)
        
        # ── Performance monitoring ─────────────────────────────
        self.frame_times = []
        self.performance_mode = "auto"  # auto, quality, speed
        self.memory_tracker = MemoryTracker()
        
        self._setup_styles()
        self._build_ui()
        
        self.root.bind("<Left>", lambda e: self.step_frame(-1))
        self.root.bind("<Right>", lambda e: self.step_frame(1))
        self.root.bind("<space>", lambda e: self.toggle_play())
        self.root.protocol("WM_DELETE_WINDOW", self._on_close)
        
        # Start memory monitor (without psutil)
        self._start_memory_monitor()
    
    def _on_close(self):
        self.merged_index.clear()
        for entry in self.file_cache.values():
            try:
                entry["mm"].close()
            except Exception:
                pass
        self.file_cache.clear()
        
        # Gracefully shut down thread pool
        self.prefetch_executor.shutdown(wait=False)
        self.root.destroy()
    
    def _start_memory_monitor(self):
        """Periodic memory usage check without psutil"""
        def monitor():
            if hasattr(self, 'root') and self.root.winfo_exists():
                try:
                    # Get approximate memory usage
                    memory_mb = self.memory_tracker.get_memory_usage()
                    
                    if memory_mb > 0 and memory_mb > 2048:  # > 2GB
                        # Memory pressure - clear caches
                        self.merged_index._frame_cache.clear()
                        gc.collect()
                        cache_stats = self.merged_index.get_cache_stats()
                        self._set_status(f"Memory: {memory_mb:.0f}MB - Cache cleared (hit rate: {cache_stats['hit_rate']:.0f}%)", color=WARNING)
                    elif memory_mb > 0:
                        cache_stats = self.merged_index.get_cache_stats()
                        self.lbl_footer_info.config(text=f"Memory: {memory_mb:.0f}MB | Cache: {cache_stats['size']} frames | Hit: {cache_stats['hit_rate']:.0f}%")
                    
                    # Schedule next check
                    self.root.after(30000, monitor)  # Check every 30 seconds
                except Exception as e:
                    # Silently handle errors in memory monitoring
                    self.root.after(30000, monitor)
        
        # Start monitor thread
        def monitor_wrapper():
            try:
                # Wait a bit for the UI to initialize
                time.sleep(5)
                if hasattr(self, 'root') and self.root.winfo_exists():
                    self.root.after(0, monitor)
            except Exception:
                pass
        
        threading.Thread(target=monitor_wrapper, daemon=True).start()
    
    def _adjust_quality_for_performance(self, frame_time_ms: float):
        """Dynamic quality adjustment based on playback performance"""
        if not self.is_playing:
            return
        
        self.frame_times.append(frame_time_ms)
        if len(self.frame_times) > 30:
            self.frame_times.pop(0)
        
        avg_time = sum(self.frame_times) / len(self.frame_times)
        target_time = 1000 / self.fps_var.get()
        
        if avg_time > target_time * 1.2 and self.decode_quality > 50:
            # Too slow, reduce quality
            self.decode_quality = max(50, self.decode_quality - 10)
            self._set_status(f"Performance: Quality reduced to {self.decode_quality}%", color=WARNING)
        elif avg_time < target_time * 0.8 and self.decode_quality < 100:
            # Fast enough, increase quality
            self.decode_quality = min(100, self.decode_quality + 5)
    
    # ── UI Setup ──
    def _setup_styles(self):
        s = ttk.Style()
        s.theme_use("clam")
        s.configure("Dark.TFrame", background=BG_DARK)
        s.configure("Panel.TFrame", background=BG_PANEL)
        s.configure("Card.TFrame", background=BG_CARD)
        s.configure("Toolbar.TFrame", background=BG_TOOLBAR)
        s.configure("TLabel", background=BG_PANEL, foreground=FG_MAIN, font=("Segoe UI", 10))
        s.configure("Dim.TLabel", background=BG_PANEL, foreground=FG_DIM, font=("Segoe UI", 9))
        s.configure("Title.TLabel", background=BG_DARK, foreground=FG_BRIGHT, font=("Segoe UI", 13, "bold"))
        s.configure("Primary.TButton", font=("Segoe UI", 10, "bold"), padding=6,
                    background=BTN_PRIMARY, foreground="white", borderwidth=0, relief="flat")
        s.map("Primary.TButton", background=[("active", BTN_HOVER)])
        s.configure("Success.TButton", font=("Segoe UI", 10, "bold"), padding=6,
                    background="#0a4d36", foreground="#00c896", borderwidth=0, relief="flat")
        s.configure("Warn.TButton", font=("Segoe UI", 10, "bold"), padding=6,
                    background="#6b3200", foreground="#f5a623", borderwidth=0, relief="flat")
        s.configure("Merge.TButton", font=("Segoe UI", 10, "bold"), padding=6,
                    background="#3a1a6b", foreground="#c084fc", borderwidth=0, relief="flat")
        s.configure("Export.TButton", font=("Segoe UI", 10, "bold"), padding=6,
                    background="#4a148c", foreground="#e1bee7", borderwidth=0, relief="flat")
        s.configure("Horizontal.TScale", background=BG_PANEL, troughcolor="#1e2540",
                    sliderlength=14, sliderrelief="flat")
        s.configure("Survey.Horizontal.TProgressbar",
                    troughcolor="#1e2540", background=ACCENT, thickness=4)
    
    def _build_ui(self):
        # Top bar
        topbar = tk.Frame(self.root, bg=BG_DARK, height=56)
        topbar.pack(fill=tk.X, side=tk.TOP)
        topbar.pack_propagate(False)
        
        logo_frame = tk.Frame(topbar, bg=BG_DARK)
        logo_frame.pack(side=tk.LEFT, padx=(14, 0), pady=8)
        tk.Label(logo_frame, text="⬡", bg=BG_DARK, fg=ACCENT, font=("Segoe UI", 20)).pack(side=tk.LEFT, padx=(0, 6))
        txt_frame = tk.Frame(logo_frame, bg=BG_DARK)
        txt_frame.pack(side=tk.LEFT)
        tk.Label(txt_frame, text="PGR SURVEY VIEWER", bg=BG_DARK, fg=FG_BRIGHT, font=("Segoe UI", 14, "bold")).pack(anchor="w")
        tk.Label(txt_frame, text="Multi-Select Merge · 12-bit · GPS Track", bg=BG_DARK, fg=FG_DIM, font=("Segoe UI", 8)).pack(anchor="w")
        
        badge_frame = tk.Frame(topbar, bg=BG_DARK)
        badge_frame.pack(side=tk.RIGHT, padx=14, pady=10)
        tk.Label(badge_frame, text="v6.3", bg=BTN_PRIMARY, fg="white", font=("Segoe UI", 8, "bold"), padx=8, pady=3).pack(side=tk.RIGHT, padx=4)
        tk.Label(badge_frame, text="ULTRASPEED", bg=ACCENT2, fg="white", font=("Segoe UI", 8, "bold"), padx=8, pady=3).pack(side=tk.RIGHT, padx=4)
        
        tk.Frame(self.root, bg=ACCENT, height=2).pack(fill=tk.X)
        
        # Main paned layout
        self.paned = ttk.PanedWindow(self.root, orient=tk.HORIZONTAL)
        self.paned.pack(fill=tk.BOTH, expand=True, padx=6, pady=6)
        
        self.left_panel = ttk.Frame(self.paned, style="Panel.TFrame")
        self.right_panel = ttk.Frame(self.paned, style="Panel.TFrame")
        self.paned.add(self.left_panel, weight=3)
        self.paned.add(self.right_panel, weight=2)
        
        self._build_left()
        self._build_right()
        self._build_footer()
    
    def _build_left(self):
        p = self.left_panel
        
        # Row 1: File loading
        row1 = tk.Frame(p, bg=BG_TOOLBAR, pady=4)
        row1.pack(fill=tk.X, padx=0, pady=(0, 2))

        # ── Data Source mode switcher (first control) ──
        mode_frame = tk.Frame(row1, bg=BG_TOOLBAR, highlightthickness=1, highlightbackground="#2a3040")
        mode_frame.pack(side=tk.LEFT, padx=(8, 8), pady=2)
        tk.Label(mode_frame, text="Data Source:", bg=BG_TOOLBAR, fg=FG_DIM, font=("Segoe UI", 8)).pack(side=tk.LEFT, padx=(6, 4))
        self.mode_var = tk.StringVar(value=self.data_mode)
        mode_combo = ttk.Combobox(mode_frame, textvariable=self.mode_var, values=["LIDAR", "LCMS"],
                                   state="readonly", width=7, font=("Segoe UI", 9, "bold"))
        mode_combo.pack(side=tk.LEFT, padx=(0, 6), pady=3)
        mode_combo.bind("<<ComboboxSelected>>", self.on_data_mode_changed)

        ttk.Button(row1, text="📂  Load PGR Folder", style="Primary.TButton", command=self.load_pgr_folder).pack(side=tk.LEFT, padx=(0, 4), pady=2)
        
        # Checkbox dropdown for file selection
        self.file_selector = CheckboxDropdown(row1, [], on_select_callback=self.on_file_selection_changed)
        self.file_selector.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=4)
        
        # Merge Selected button
        self.merge_selected_btn = tk.Button(
            row1, 
            text="🔗 Merge Selected Files", 
            bg=MERGED_COLOR, 
            fg="white", 
            font=("Segoe UI", 9, "bold"),
            relief="flat",
            cursor="hand2",
            padx=8,
            pady=2,
            command=self.merge_selected_files
        )
        self.merge_selected_btn.pack(side=tk.LEFT, padx=4, pady=2)
        
        self.gps_load_btn = ttk.Button(row1, style="Success.TButton", command=self._gps_load_dispatch)
        self.gps_load_btn.pack(side=tk.LEFT, padx=4, pady=2)
        self._refresh_gps_button()
        ttk.Button(row1, text="📊  Export Excel", style="Warn.TButton", command=self.export_to_excel).pack(side=tk.LEFT, padx=(0, 4), pady=2)
        ttk.Button(row1, text="📸 Export Cam0+GPS", style="Export.TButton", command=self.export_cam0_and_excel).pack(side=tk.LEFT, padx=(0, 8), pady=2)
        ttk.Button(row1, text="🎬 Export Video", style="Merge.TButton", command=self.export_video_dialog).pack(side=tk.LEFT, padx=(0, 8), pady=2)
        
        # Row 2: View & camera settings
        row2 = tk.Frame(p, bg=BG_PANEL)
        row2.pack(fill=tk.X, padx=6, pady=2)
        
        tk.Label(row2, text="View:", bg=BG_PANEL, fg=FG_DIM, font=("Segoe UI", 9)).pack(side=tk.LEFT)
        self.view_var = tk.StringVar(value="Single Camera")
        view_combo = ttk.Combobox(row2, textvariable=self.view_var, values=list(DUAL_VIEW_OPTIONS.keys()), state="readonly", width=20, font=("Segoe UI", 9))
        view_combo.pack(side=tk.LEFT, padx=6)
        view_combo.bind("<<ComboboxSelected>>", self.on_view_changed)
        
        tk.Label(row2, text="Camera:", bg=BG_PANEL, fg=FG_DIM, font=("Segoe UI", 9)).pack(side=tk.LEFT, padx=(10, 0))
        self.cam_var = tk.IntVar(value=2)
        cam_spin = tk.Spinbox(row2, from_=0, to=5, width=3, textvariable=self.cam_var, font=("Courier New", 9), bg=BG_CARD, fg=ACCENT, relief="flat", command=self.on_camera_changed)
        cam_spin.pack(side=tk.LEFT, padx=6)
        self.cam_label = tk.Label(row2, text="Camera 2 – Right", bg=BG_PANEL, fg=FG_DIM, font=("Segoe UI", 9))
        self.cam_label.pack(side=tk.LEFT, padx=5)
        
        # Zoom controls
        zoom_group = tk.Frame(row2, bg=BG_PANEL)
        zoom_group.pack(side=tk.LEFT, padx=(14, 0))
        tk.Label(zoom_group, text="Zoom:", bg=BG_PANEL, fg=FG_DIM, font=("Segoe UI", 9)).pack(side=tk.LEFT)
        tk.Button(zoom_group, text="－", width=2, bg=BG_CARD, fg=FG_MAIN, font=("Segoe UI", 9, "bold"),
                  relief="flat", cursor="hand2", command=self.zoom_out).pack(side=tk.LEFT, padx=(6, 1))
        self.zoom_pct_label = tk.Label(zoom_group, text="100%", bg=BG_PANEL, fg=ACCENT, font=("Courier New", 9, "bold"), width=5)
        self.zoom_pct_label.pack(side=tk.LEFT, padx=2)
        tk.Button(zoom_group, text="＋", width=2, bg=BG_CARD, fg=FG_MAIN, font=("Segoe UI", 9, "bold"),
                  relief="flat", cursor="hand2", command=self.zoom_in).pack(side=tk.LEFT, padx=(1, 6))
        tk.Button(zoom_group, text="Fit", bg=BG_CARD, fg=FG_DIM, font=("Segoe UI", 8), relief="flat",
                  padx=6, cursor="hand2", command=self.reset_zoom).pack(side=tk.LEFT)
        
        tk.Label(row2, text="Map:", bg=BG_PANEL, fg=FG_DIM, font=("Segoe UI", 9)).pack(side=tk.LEFT, padx=(14, 0))
        self.map_type_var = tk.StringVar(value="Satellite (Google)")
        map_combo = ttk.Combobox(row2, textvariable=self.map_type_var, values=list(MAP_TILES.keys()), state="readonly", width=18, font=("Segoe UI", 9))
        map_combo.pack(side=tk.LEFT, padx=6)
        map_combo.bind("<<ComboboxSelected>>", self.on_map_type_changed)
        
        self.lbl_status = tk.Label(row2, text="⬤  Ready", bg=BG_PANEL, fg=SUCCESS, font=("Segoe UI", 9, "bold"))
        self.lbl_status.pack(side=tk.RIGHT, padx=8)
        
        # Row 3: Color gain sliders
        row3 = tk.Frame(p, bg=BG_PANEL)
        row3.pack(fill=tk.X, padx=6, pady=(2, 4))
        
        tk.Label(row3, text="Color Balance:", bg=BG_PANEL, fg=FG_DIM, font=("Segoe UI", 9)).pack(side=tk.LEFT, padx=(0, 10))
        
        def mk_slider(parent, label, var, color):
            f = tk.Frame(parent, bg=BG_PANEL)
            f.pack(side=tk.LEFT, padx=8)
            tk.Label(f, text=label, fg=color, bg=BG_PANEL, font=("Segoe UI", 9, "bold")).pack(side=tk.LEFT)
            slider = tk.Scale(f, from_=0.1, to=3.5, resolution=0.05, orient=tk.HORIZONTAL, variable=var, bg=BG_PANEL, fg=FG_MAIN, highlightthickness=0, length=100, showvalue=False, command=lambda x: self.on_gain_changed())
            slider.pack(side=tk.LEFT, padx=(4, 0))
            tk.Label(f, textvariable=var, bg=BG_PANEL, fg=ACCENT, font=("Courier New", 9), width=4).pack(side=tk.LEFT)
        
        self.r_var = tk.DoubleVar(value=1.1)
        self.g_var = tk.DoubleVar(value=1.0)
        self.b_var = tk.DoubleVar(value=1.1)
        mk_slider(row3, "R", self.r_var, "#ff5252")
        mk_slider(row3, "G", self.g_var, "#69f0ae")
        mk_slider(row3, "B", self.b_var, "#40c4ff")
        
        tk.Button(row3, text="Reset", bg=BG_CARD, fg=FG_DIM, font=("Segoe UI", 8), relief="flat", padx=8, cursor="hand2", command=self.reset_gains).pack(side=tk.LEFT, padx=10)
        
        self.merged_badge = tk.Label(row3, text="⬡ MERGED MODE — Selected PGR files", bg=MERGED_COLOR, fg="white", font=("Segoe UI", 8, "bold"), padx=8, pady=2)
        self.merged_badge.pack_forget()  # Initially hidden
        
        # Video canvas
        canvas_outer = tk.Frame(p, bg="#080b10", bd=1, relief="flat", highlightthickness=1, highlightbackground="#1e2540")
        canvas_outer.pack(fill=tk.BOTH, expand=True, padx=6, pady=4)
        self.video_canvas = tk.Canvas(canvas_outer, bg="#000000", highlightthickness=0)
        self.video_canvas.pack(fill=tk.BOTH, expand=True)
        
        self.canvas_info_label = tk.Label(self.video_canvas, text="No file loaded", bg="#0b0e14", fg=FG_DIM, font=("Segoe UI", 10))
        self.canvas_info_label.place(x=10, y=10)
        self.canvas_frame_label = tk.Label(self.video_canvas, text="", bg="#0b0e14", fg=FG_DIM, font=("Courier New", 9))
        self.canvas_frame_label.place(relx=1.0, x=-10, y=10, anchor="ne")
        self.canvas_file_label = tk.Label(self.video_canvas, text="", bg="#0b0e14", fg=MERGED_COLOR, font=("Segoe UI", 8, "bold"))
        self.canvas_file_label.place(x=10, rely=1.0, y=-26)
        self.canvas_zoom_label = tk.Label(self.video_canvas, text="", bg="#0b0e14", fg=ACCENT, font=("Courier New", 9, "bold"))
        self.canvas_zoom_label.place(relx=1.0, rely=1.0, x=-10, y=-26, anchor="se")
        self.canvas_zoom_hint = tk.Label(self.video_canvas, text="Scroll: zoom  ·  Drag: pan  ·  Double-click: fit",
                                          bg="#0b0e14", fg=FG_DIM, font=("Segoe UI", 8))
        self.canvas_zoom_hint.place(relx=0.5, rely=1.0, y=-8, anchor="s")
        
        # ── Zoom / Pan bindings on the GR view canvas ───────────
        self.video_canvas.bind("<MouseWheel>", self._on_mousewheel_zoom)      # Windows / macOS
        self.video_canvas.bind("<Button-4>", self._on_mousewheel_zoom)        # Linux scroll up
        self.video_canvas.bind("<Button-5>", self._on_mousewheel_zoom)        # Linux scroll down
        self.video_canvas.bind("<ButtonPress-1>", self._on_pan_start)
        self.video_canvas.bind("<B1-Motion>", self._on_pan_drag)
        self.video_canvas.bind("<ButtonRelease-1>", self._on_pan_end)
        self.video_canvas.bind("<Double-Button-1>", lambda e: self.reset_zoom())
        self.video_canvas.bind("<Configure>", self._on_canvas_resize)
        self.root.bind("<plus>", lambda e: self.zoom_in())
        self.root.bind("<equal>", lambda e: self.zoom_in())
        self.root.bind("<minus>", lambda e: self.zoom_out())
        self.root.bind("<KP_Add>", lambda e: self.zoom_in())
        self.root.bind("<KP_Subtract>", lambda e: self.zoom_out())
        
        # Timeline slider
        tl_frame = tk.Frame(p, bg=BG_PANEL)
        tl_frame.pack(fill=tk.X, padx=6, pady=(0, 2))
        tk.Label(tl_frame, text="0", bg=BG_PANEL, fg=FG_DIM, font=("Segoe UI", 8)).pack(side=tk.LEFT)
        self.slider_var = tk.IntVar()
        self.slider = ttk.Scale(tl_frame, from_=0, to=100, orient=tk.HORIZONTAL, variable=self.slider_var, command=self.on_slider_move, style="Horizontal.TScale")
        self.slider.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=6)
        self.lbl_frame_max = tk.Label(tl_frame, text="0", bg=BG_PANEL, fg=FG_DIM, font=("Segoe UI", 8))
        self.lbl_frame_max.pack(side=tk.LEFT)
        
        # Playback controls
        ctrl = tk.Frame(p, bg=BG_CARD, pady=4)
        ctrl.pack(fill=tk.X, padx=6, pady=(2, 6))
        
        self.play_btn = tk.Button(ctrl, text="▶", bg=BTN_PRIMARY, fg="white", font=("Segoe UI", 14, "bold"), relief="flat", bd=0, padx=14, pady=4, cursor="hand2", command=self.toggle_play)
        self.play_btn.pack(side=tk.LEFT, padx=6, pady=4)
        
        for text, step in [("⏮–10", -10), ("◀–1", -1), ("+1▶", 1), ("+10⏭", 10)]:
            tk.Button(ctrl, text=text, bg=BG_PANEL, fg=FG_MAIN, font=("Segoe UI", 9), relief="flat", bd=0, padx=10, pady=4, cursor="hand2", command=lambda s=step: self.step_frame(s)).pack(side=tk.LEFT, padx=2)
        
        tk.Label(ctrl, text="FPS:", bg=BG_CARD, fg=FG_DIM, font=("Segoe UI", 9)).pack(side=tk.LEFT, padx=(16, 0))
        self.fps_var = tk.IntVar(value=15)
        ttk.Scale(ctrl, from_=1, to=60, orient=tk.HORIZONTAL, variable=self.fps_var, length=100, style="Horizontal.TScale").pack(side=tk.LEFT, padx=4)
        self.lbl_fps = tk.Label(ctrl, text="15", bg=BG_CARD, fg=ACCENT, font=("Segoe UI", 9, "bold"), width=3)
        self.lbl_fps.pack(side=tk.LEFT)
        self.fps_var.trace_add("write", lambda *_: self.lbl_fps.config(text=str(self.fps_var.get())))
        
        tk.Button(ctrl, text="💾 Save GPS", bg="#003d2e", fg="#00c896", font=("Segoe UI", 9), relief="flat", bd=0, padx=12, pady=4, cursor="hand2", command=self.save_gps_to_files).pack(side=tk.RIGHT, padx=6)
        
        self.lbl_frame_cur = tk.Label(ctrl, text="Frame: — / —", bg=BG_CARD, fg=FG_DIM, font=("Courier New", 9))
        self.lbl_frame_cur.pack(side=tk.RIGHT, padx=10)
        
        # Progress bar
        self.progress_frame = tk.Frame(p, bg=BG_PANEL, pady=4, padx=6)
        self.progress_var = tk.DoubleVar(value=0)
        self.progress_bar = ttk.Progressbar(self.progress_frame, variable=self.progress_var, maximum=100, length=400, mode="determinate", style="Survey.Horizontal.TProgressbar")
        self.lbl_progress = tk.Label(self.progress_frame, text="", bg=BG_PANEL, fg=ACCENT, font=("Courier New", 8))
    
    def _build_right(self):
        p = self.right_panel
        header = tk.Frame(p, bg=BG_PANEL)
        header.pack(fill=tk.X, padx=6, pady=(6, 2))
        tk.Label(header, text="🛰  LIVE GPS MAP", bg=BG_PANEL, fg=ACCENT, font=("Segoe UI", 11, "bold")).pack(side=tk.LEFT)
        tk.Label(header, text="Left-click path → Jump video to location", bg=BG_PANEL, fg=FG_DIM, font=("Segoe UI", 8, "italic")).pack(side=tk.LEFT, padx=8)
        self.lbl_gps_coord = tk.Label(header, text="GPS: —", bg=BG_PANEL, fg=FG_DIM, font=("Courier New", 8))
        self.lbl_gps_coord.pack(side=tk.RIGHT, padx=6)
        self.lbl_gps_file = tk.Label(header, text="", bg=BG_PANEL, fg=SUCCESS, font=("Segoe UI", 8))
        self.lbl_gps_file.pack(side=tk.RIGHT, padx=8)
        
        map_frame = tk.Frame(p, bg="#000000", bd=0)
        map_frame.pack(fill=tk.BOTH, expand=True, padx=6, pady=(2, 6))
        
        if HAVE_MAP:
            self.map_widget = tkintermapview.TkinterMapView(map_frame, corner_radius=0)
            self.map_widget.pack(fill=tk.BOTH, expand=True)
            self.map_widget.set_tile_server(MAP_TILES["Satellite (Google)"], max_zoom=22)
            self.map_widget.add_left_click_map_command(self.sync_video_from_map)
        else:
            self.map_widget = None
            tk.Label(map_frame, text="tkintermapview not installed\n\npip install tkintermapview", bg="#000000", fg=FG_DIM, font=("Segoe UI", 10)).pack(expand=True)
    
    def _build_footer(self):
        foot = tk.Frame(self.root, bg="#090c12", height=26)
        foot.pack(fill=tk.X, side=tk.BOTTOM)
        foot.pack_propagate(False)
        tk.Frame(foot, bg=ACCENT, width=4).pack(side=tk.LEFT, fill=tk.Y)
        tk.Label(foot, text="  Eng. Samir Abozahra  ·  Ladybug6 12-bit  ·  Multi-Select Merge  ·  samirabozahra3@gmail.com", bg="#090c12", fg=FG_DIM, font=("Segoe UI", 8)).pack(side=tk.LEFT, pady=4)
        self.lbl_footer_info = tk.Label(foot, text="", bg="#090c12", fg=FG_DIM, font=("Segoe UI", 8))
        self.lbl_footer_info.pack(side=tk.RIGHT, padx=12)
    
    # ── File Loading (optimized multi-threaded) ──────────────────────────────────
    def load_pgr_folder(self):
        if self.is_indexing:
            return
        folder = filedialog.askdirectory(title="Select Folder Containing *000000.pgr Files")
        if not folder:
            return
        
        for entry in self.file_cache.values():
            try:
                entry["mm"].close()
            except Exception:
                pass
        self.file_cache.clear()
        self.merged_index.clear()
        
        self.pgr_files = sorted(Path(folder).glob("*000000.pgr"))
        if not self.pgr_files:
            messagebox.showwarning("No Files", "No *000000.pgr files found in selected folder.")
            return
        
        # Update checkbox dropdown with file names
        file_names = [f.name for f in self.pgr_files]
        self.file_selector.set_values(file_names)
        self.file_selector.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=4)
        
        # Pre-load all files in background (parallel indexing)
        self._preload_all_files()
        
        self._set_status(f"Loaded {len(self.pgr_files)} PGR files. Select files to merge.", color=SUCCESS)
    
    def _preload_all_files(self):
        """Pre-load all PGR files in background using a ThreadPool for massive speedups"""
        if self.is_indexing:
            return
        
        files_to_load = [fp for fp in self.pgr_files if fp not in self.file_cache]
        if not files_to_load:
            return
        
        self.is_indexing = True
        self._show_progress(True)
        self._set_progress(0, f"Pre-loading {len(files_to_load)} files...")
        
        def worker():
            def process_file(fp):
                try:
                    mm, frames = scan_pgr_frames_optimized(fp)
                    return fp, mm, frames, None
                except Exception as e:
                    return fp, None, None, e

            loaded = 0
            total = len(files_to_load)
            
            # Use thread pool to index multiple files simultaneously
            max_workers = min(8, os.cpu_count() or 4)
            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                futures = [executor.submit(process_file, fp) for fp in files_to_load]
                for future in as_completed(futures):
                    fp, mm, frames, err = future.result()
                    if err is None:
                        self.file_cache[fp] = {"mm": mm, "frames": frames}
                    loaded += 1
                    pct = (loaded / total) * 100
                    self.root.after(0, self._set_progress, pct, f"Loaded {loaded}/{total}: {fp.name}")
            
            self.root.after(0, self._preload_done)
        
        threading.Thread(target=worker, daemon=True).start()
    
    def _preload_done(self):
        self.is_indexing = False
        self._show_progress(False)
        self._set_status(f"Pre-loaded {len(self.file_cache)} PGR files. Select files to merge.", color=SUCCESS)
    
    def on_file_selection_changed(self, selected_files):
        """Callback when file selection changes"""
        self.selected_merge_files = selected_files
        count = len(selected_files)
        if count == 0:
            self._set_status("No files selected. Select at least one file to view.", color=WARNING)
        else:
            self._set_status(f"{count} file(s) selected. Click 'Merge Selected Files' to combine.", color=ACCENT)
    
    def merge_selected_files(self):
        """Merge only the selected files"""
        if not self.selected_merge_files:
            messagebox.showwarning("No Selection", "Please select at least one PGR file to merge.")
            return
        
        # Get full paths for selected files
        selected_paths = []
        for fname in self.selected_merge_files:
            for fp in self.pgr_files:
                if fp.name == fname:
                    selected_paths.append(fp)
                    break
        
        if not selected_paths:
            messagebox.showwarning("Error", "Selected files not found.")
            return
        
        # Start merge process
        self._start_merged_indexing_selected(selected_paths)
    
    def _start_merged_indexing_selected(self, selected_paths):
        """Merge only selected files using multi-threading if missing"""
        self._stop_playback_if_running()
        self.merged_mode = True
        
        # Ensure all selected files are loaded
        files_to_load = [fp for fp in selected_paths if fp not in self.file_cache]
        
        if files_to_load:
            self.is_indexing = True
            self._show_progress(True)
            self._set_progress(0, f"Loading {len(files_to_load)} selected files...")
            
            def worker():
                def process_file(fp):
                    try:
                        mm, frames = scan_pgr_frames_optimized(fp)
                        return fp, mm, frames, None
                    except Exception as e:
                        return fp, None, None, e

                max_workers = min(8, os.cpu_count() or 4)
                with ThreadPoolExecutor(max_workers=max_workers) as executor:
                    futures = [executor.submit(process_file, fp) for fp in files_to_load]
                    for future in as_completed(futures):
                        fp, mm, frames, err = future.result()
                        if err is None:
                            self.file_cache[fp] = {"mm": mm, "frames": frames}
                
                self.root.after(0, self._do_merge_selected, selected_paths)
            
            threading.Thread(target=worker, daemon=True).start()
        else:
            self._do_merge_selected(selected_paths)
    
    def _do_merge_selected(self, selected_paths):
        """Perform the actual merge of selected files"""
        self.merged_index.clear()
        total_mb = 0
        valid_files = []
        
        for fp in selected_paths:
            cache_entry = self.file_cache.get(fp)
            if cache_entry and cache_entry["frames"]:
                self.merged_index.add_file(fp, cache_entry["mm"], cache_entry["frames"])
                total_mb += fp.stat().st_size / 1e6
                valid_files.append(fp.name)
        
        if self.merged_index.total_frames == 0:
            self.is_indexing = False
            self._show_progress(False)
            self._draw_error("No valid frames found in selected files.")
            return
        
        self.current_frame_idx = 0
        self.total_frames = self.merged_index.total_frames
        
        self.slider.config(to=max(1, self.total_frames - 1))
        self.lbl_frame_max.config(text=str(self.total_frames - 1))
        
        # Show merge badge and info (distinguish single-file vs true multi-file merge)
        if len(valid_files) == 1:
            self.merged_badge.config(text=f"⬡ SINGLE FILE — {valid_files[0]}")
        else:
            self.merged_badge.config(text=f"⬡ MERGED MODE — {len(valid_files)} files: {', '.join(valid_files[:3])}{'...' if len(valid_files) > 3 else ''}")
        self.merged_badge.pack(side=tk.RIGHT, padx=8)
        
        # Recompute which GPS run(s) belong to THIS selection only
        if self.gps_by_run:
            self._refresh_active_gps_selection()
        
        self.is_indexing = False
        self._show_progress(False)
        
        self._set_status(f"Loaded: {len(valid_files)} file(s) · {self.total_frames} total frames", color=MERGED_COLOR)
        self.lbl_footer_info.config(text=f"Loaded {len(valid_files)} file(s) · {total_mb:.1f} MB total")
        self.show_frame()
    
    # ── Frame rendering with caching & background prefetch ───────
    def show_frame(self):
        frame_start = time.time()
        
        # Generate universal cache key
        cache_key = self.merged_index.get_cache_key(
            self.current_frame_idx, self.view_mode, 
            self.current_cam_idx, (self.r_gain, self.g_gain, self.b_gain, self.decode_quality)
        ) if self.merged_mode else f"single_{self.current_frame_idx}_{self.view_mode}_{self.current_cam_idx}_{self.decode_quality}"
        
        # Universal cache check
        cached_img = self.merged_index.get_cached_frame(cache_key)
        
        if cached_img:
            self._current_pil_image = cached_img
            if self.merged_mode:
                fname = self.merged_index.file_name_at(self.current_frame_idx)
                self.canvas_file_label.config(text=f"⬡ {fname}")
            
            # Setup info label natively
            if self.view_mode == "dual": info_text = "Dual Panoramic (Cam 2+3)"
            elif self.view_mode == "grid": info_text = "All 6 Cameras Grid"
            else: info_text = CAMERA_NAMES.get(self.current_cam_idx, f"Camera {self.current_cam_idx}")
            self.canvas_info_label.config(text=info_text)

            self._update_display_info()
            self._render_to_canvas()
            
            # Trigger background prefetching for butter smooth playback
            if self.is_playing:
                self._prefetch_upcoming()
                self._adjust_quality_for_performance((time.time() - frame_start) * 1000)
            return
        
        # Cache Miss - Decode actively
        if self.merged_mode:
            if self.merged_index.total_frames == 0:
                return
            mm, frame_data = self.merged_index.get_frame_data(self.current_frame_idx)
            fname = self.merged_index.file_name_at(self.current_frame_idx)
            self.canvas_file_label.config(text=f"⬡ {fname}")
        else:
            if not self._single_mm or not self._single_frames:
                return
            mm = self._single_mm
            frame_data = self._single_frames[self.current_frame_idx]
            self.canvas_file_label.config(text="")
        
        try:
            if self.view_mode == "dual":
                img = decode_dual_panoramic_optimized(mm, frame_data["planes"],
                                                       self.r_gain, self.g_gain, self.b_gain,
                                                       decode_quality=self.decode_quality)
                info_text = "Dual Panoramic (Cam 2+3)"
            elif self.view_mode == "grid":
                img = decode_all_cameras_grid_optimized(mm, frame_data["planes"],
                                                         self.r_gain, self.g_gain, self.b_gain,
                                                         decode_quality=self.decode_quality)
                info_text = "All 6 Cameras Grid"
            else:
                img = decode_camera_frame_optimized(mm, frame_data["planes"],
                                                     self.current_cam_idx,
                                                     self.r_gain, self.g_gain, self.b_gain,
                                                     decode_quality=self.decode_quality)
                info_text = CAMERA_NAMES.get(self.current_cam_idx, f"Camera {self.current_cam_idx}")
            
            # Populate cache globally
            self.merged_index.cache_frame(cache_key, img)
            
            self._current_pil_image = img
            self.canvas_info_label.config(text=info_text)
            self._update_display_info()
            self._render_to_canvas()
            
            if self.is_playing:
                self._prefetch_upcoming()
            
            # Performance monitoring
            frame_time = (time.time() - frame_start) * 1000
            self._adjust_quality_for_performance(frame_time)
            
        except Exception as e:
            self._draw_error(f"Frame decode error:\n{e}")

    def _prefetch_upcoming(self):
        """Asynchronously decode and cache the next 4 frames to keep playback buttery smooth."""
        for offset in range(1, 5):
            next_idx = self.current_frame_idx + offset
            if next_idx >= self.total_frames:
                break
            self.prefetch_executor.submit(
                self._decode_and_cache_bg, 
                next_idx, self.view_mode, self.current_cam_idx,
                self.r_gain, self.g_gain, self.b_gain, 
                self.decode_quality, self.merged_mode
            )

    def _decode_and_cache_bg(self, idx, view_mode, cam_idx, r_gain, g_gain, b_gain, decode_quality, merged_mode):
        """The threaded background worker for decoding future frames directly into the LRU Cache"""
        cache_key = self.merged_index.get_cache_key(
            idx, view_mode, cam_idx, (r_gain, g_gain, b_gain, decode_quality)
        ) if merged_mode else f"single_{idx}_{view_mode}_{cam_idx}_{decode_quality}"

        if self.merged_index.get_cached_frame(cache_key) is not None:
            return  # Already cached

        try:
            if merged_mode:
                if self.merged_index.total_frames == 0: return
                mm, frame_data = self.merged_index.get_frame_data(idx)
            else:
                if not self._single_mm or not self._single_frames: return
                mm = self._single_mm
                frame_data = self._single_frames[idx]

            if view_mode == "dual":
                img = decode_dual_panoramic_optimized(mm, frame_data["planes"], r_gain, g_gain, b_gain, decode_quality=decode_quality)
            elif view_mode == "grid":
                img = decode_all_cameras_grid_optimized(mm, frame_data["planes"], r_gain, g_gain, b_gain, decode_quality=decode_quality)
            else:
                img = decode_camera_frame_optimized(mm, frame_data["planes"], cam_idx, r_gain, g_gain, b_gain, decode_quality=decode_quality)

            self.merged_index.cache_frame(cache_key, img)
        except Exception:
            pass
    
    def _update_display_info(self):
        self.canvas_frame_label.config(text=f"Frame {self.current_frame_idx:06d} / {self.total_frames - 1:06d}")
        self.slider_var.set(self.current_frame_idx)
        self.lbl_frame_cur.config(text=f"Frame: {self.current_frame_idx} / {self.total_frames - 1}")
        self._update_gps_marker()
    
    def _render_to_canvas(self):
        if self._current_pil_image is None:
            return
        cw = self.video_canvas.winfo_width()
        ch = self.video_canvas.winfo_height()
        if cw < 10 or ch < 10:
            cw, ch = 800, 600
        iw, ih = self._current_pil_image.size
        
        # base_scale = "fit to window" scale (the original, unzoomed view)
        base_scale = min(cw / iw, ch / ih)
        scale = base_scale * self.zoom_level
        nw, nh = max(1, int(round(iw * scale))), max(1, int(round(ih * scale)))
        
        # Keep pan within sane bounds for the current size/zoom
        self._clamp_pan(cw, ch, nw, nh)
        
        # ⚡ Adaptive Resampling Logic (Massive Performance Boost) ⚡
        # NEAREST provides flawless 60FPS dragging.
        # BILINEAR provides lightweight video playback.
        # LANCZOS provides razor-sharp images when perfectly still.
        if self._is_panning:
            resample_filter = Image.NEAREST
        elif self.is_playing:
            resample_filter = Image.BILINEAR
        else:
            resample_filter = Image.LANCZOS

        resized = self._current_pil_image.resize((nw, nh), resample_filter)
        self._photo_ref = ImageTk.PhotoImage(resized)
        
        self.video_canvas.delete("all")
        cx = cw // 2 + int(round(self.pan_offset_x))
        cy = ch // 2 + int(round(self.pan_offset_y))
        self.video_canvas.create_image(cx, cy, anchor="center", image=self._photo_ref)
        self._update_zoom_label()
    
    # ── Zoom / Pan (GR view panel) ──────────────────────────────
    def _update_zoom_label(self):
        if hasattr(self, "zoom_pct_label"):
            pct = int(round(self.zoom_level * 100))
            self.zoom_pct_label.config(text=f"{pct}%")
        if hasattr(self, "canvas_zoom_label"):
            pct = int(round(self.zoom_level * 100))
            self.canvas_zoom_label.config(text=f"🔍 {pct}%" if pct != 100 else "")
    
    def _clamp_pan(self, cw, ch, nw, nh):
        """Keep the panned image always overlapping the visible canvas."""
        max_x = max(0.0, (nw - cw) / 2)
        max_y = max(0.0, (nh - ch) / 2)
        self.pan_offset_x = max(-max_x, min(max_x, self.pan_offset_x))
        self.pan_offset_y = max(-max_y, min(max_y, self.pan_offset_y))
    
    def _zoom_at_point(self, mx, my, new_zoom):
        """Zoom in/out while keeping the image point under (mx, my) fixed on screen,
        so the same view stays centered under the cursor — exactly like a map viewer."""
        new_zoom = max(self.zoom_min, min(self.zoom_max, new_zoom))
        if self._current_pil_image is None:
            self.zoom_level = new_zoom
            self._update_zoom_label()
            return
        
        cw = self.video_canvas.winfo_width()
        ch = self.video_canvas.winfo_height()
        if cw < 10 or ch < 10:
            self.zoom_level = new_zoom
            self._render_to_canvas()
            return
        
        iw, ih = self._current_pil_image.size
        base_scale = min(cw / iw, ch / ih)
        old_scale = base_scale * self.zoom_level
        new_scale = base_scale * new_zoom
        
        if old_scale <= 0:
            old_scale = base_scale
        
        # Current on-screen top-left of the rendered image
        cx_old = cw / 2 + self.pan_offset_x
        cy_old = ch / 2 + self.pan_offset_y
        old_nw = iw * old_scale
        old_nh = ih * old_scale
        left_old = cx_old - old_nw / 2
        top_old = cy_old - old_nh / 2
        
        # Image-space coordinate currently under the cursor
        ix = (mx - left_old) / old_scale
        iy = (my - top_old) / old_scale
        
        # Solve new pan so that same image point stays under the cursor
        new_nw = iw * new_scale
        new_nh = ih * new_scale
        new_cx = mx - ix * new_scale + new_nw / 2
        new_cy = my - iy * new_scale + new_nh / 2
        
        self.zoom_level = new_zoom
        self.pan_offset_x = new_cx - cw / 2
        self.pan_offset_y = new_cy - ch / 2
        self._clamp_pan(cw, ch, new_nw, new_nh)
        self._render_to_canvas()
    
    def _on_mousewheel_zoom(self, event):
        if self._current_pil_image is None:
            return
        direction = 0
        if getattr(event, "num", None) == 4:
            direction = 1
        elif getattr(event, "num", None) == 5:
            direction = -1
        elif getattr(event, "delta", 0) > 0:
            direction = 1
        elif getattr(event, "delta", 0) < 0:
            direction = -1
        if direction == 0:
            return
        if direction > 0:
            new_zoom = min(self.zoom_max, self.zoom_level * self.zoom_step)
        else:
            new_zoom = max(self.zoom_min, self.zoom_level / self.zoom_step)
        self._zoom_at_point(event.x, event.y, new_zoom)
        return "break"
    
    def zoom_in(self):
        cw = self.video_canvas.winfo_width() or 800
        ch = self.video_canvas.winfo_height() or 600
        self._zoom_at_point(cw / 2, ch / 2, min(self.zoom_max, self.zoom_level * self.zoom_step))
    
    def zoom_out(self):
        cw = self.video_canvas.winfo_width() or 800
        ch = self.video_canvas.winfo_height() or 600
        self._zoom_at_point(cw / 2, ch / 2, max(self.zoom_min, self.zoom_level / self.zoom_step))
    
    def reset_zoom(self):
        """Return to the default fit-to-window view (same view, zoomed out fully)."""
        self.zoom_level = 1.0
        self.pan_offset_x = 0.0
        self.pan_offset_y = 0.0
        self._render_to_canvas()
    
    def _on_pan_start(self, event):
        if self._current_pil_image is None:
            return
        self._is_panning = True
        self._pan_start = (event.x, event.y, self.pan_offset_x, self.pan_offset_y)
        self.video_canvas.config(cursor="fleur")
    
    def _on_pan_drag(self, event):
        if not self._is_panning or self._pan_start is None or self._current_pil_image is None:
            return
        sx, sy, ox, oy = self._pan_start
        self.pan_offset_x = ox + (event.x - sx)
        self.pan_offset_y = oy + (event.y - sy)
        self._render_to_canvas()
    
    def _on_pan_end(self, event=None):
        self._is_panning = False
        self._pan_start = None
        if hasattr(self, "video_canvas"):
            self.video_canvas.config(cursor="")
        # Force a high-quality redraw (Lanczos) now that dragging has stopped
        self._render_to_canvas()
    
    def _on_canvas_resize(self, event=None):
        # Debounce rapid resize events, then redraw the current frame at the new size
        if self._resize_after_id:
            try:
                self.root.after_cancel(self._resize_after_id)
            except Exception:
                pass
        self._resize_after_id = self.root.after(60, self._render_to_canvas)
    
    def _draw_placeholder(self, text: str):
        self.video_canvas.delete("all")
        w = self.video_canvas.winfo_width() or 800
        h = self.video_canvas.winfo_height() or 600
        self.video_canvas.create_text(w // 2, h // 2, text=text, fill=FG_DIM, font=("Segoe UI", 12), justify="center")
    
    def _draw_error(self, text: str):
        self.video_canvas.delete("all")
        w = self.video_canvas.winfo_width() or 800
        h = self.video_canvas.winfo_height() or 600
        self.video_canvas.create_text(w // 2, h // 2, text=f"⚠  {text}", fill=DANGER, font=("Segoe UI", 12), justify="center")
    
    def _show_progress(self, visible: bool):
        if visible:
            self.progress_frame.pack(fill=tk.X, padx=6, pady=4)
            self.progress_bar.pack(side=tk.LEFT, fill=tk.X, expand=True)
            self.lbl_progress.pack(side=tk.LEFT, padx=6)
        else:
            self.progress_frame.pack_forget()
            self.progress_var.set(0)
    
    # ── View / Camera / Gain ──────────────────────────────────
    def on_view_changed(self, _event=None):
        self.view_mode = DUAL_VIEW_OPTIONS.get(self.view_var.get(), "single")
        # Aspect ratio differs between single/dual/grid views, so start fresh at fit-to-window
        self.reset_zoom()
        self.show_frame()
    
    def on_camera_changed(self):
        self.current_cam_idx = self.cam_var.get()
        self.cam_label.config(text=CAMERA_NAMES.get(self.current_cam_idx, f"Camera {self.current_cam_idx}"))
        if self.view_mode == "single":
            self.show_frame()
    
    def on_gain_changed(self):
        self.r_gain = self.r_var.get()
        self.g_gain = self.g_var.get()
        self.b_gain = self.b_var.get()
        self.show_frame()
    
    def reset_gains(self):
        self.r_var.set(1.1)
        self.g_var.set(1.0)
        self.b_var.set(1.1)
        self.on_gain_changed()
    
    # ── Playback ──────────────────────────────────────────────
    def _stop_playback_if_running(self):
        if self.is_playing:
            self.toggle_play()
    
    def toggle_play(self):
        if (self.merged_mode and self.merged_index.total_frames == 0) or \
           (not self.merged_mode and not self._single_frames):
            return
        self.is_playing = not self.is_playing
        if self.is_playing:
            self.play_btn.config(text="⏸", bg="#6a1fa2")
            self._play_loop()
        else:
            self.play_btn.config(text="▶", bg=BTN_PRIMARY)
            if self._after_id:
                self.root.after_cancel(self._after_id)
                self._after_id = None
            # Restore full decode quality for a crisp paused / zoomed-in view
            if self.decode_quality < 100:
                self.decode_quality = 100
                self.frame_times = []
            # Forcing a show_frame instantly brings back Lanczos upscaling on stop
            self.show_frame()
    
    def _play_loop(self):
        if not self.is_playing:
            return
        if self.current_frame_idx < self.total_frames - 1:
            self.current_frame_idx += 1
            self.show_frame()
            interval = max(10, int(1000 / max(1, self.fps_var.get())))
            self._after_id = self.root.after(interval, self._play_loop)
        else:
            self.toggle_play()
            self._set_status("Playback complete", color=WARNING)
    
    def step_frame(self, step: int):
        if self.total_frames == 0:
            return
        new_idx = max(0, min(self.current_frame_idx + step, self.total_frames - 1))
        if new_idx != self.current_frame_idx:
            if self.is_playing:
                self.toggle_play()
            self.current_frame_idx = new_idx
            self.show_frame()
    
    def on_slider_move(self, val):
        if self.total_frames == 0:
            return
        new_idx = int(float(val))
        if new_idx != self.current_frame_idx:
            if self.is_playing:
                self.toggle_play()
            self.current_frame_idx = new_idx
            self.show_frame()
    
    # ── GPS (optimized for large files) ─────────────────────────
    # ── Data source mode (LIDAR / LCMS) ──────────────────────
    def _refresh_gps_button(self):
        """Update the single GPS-load toolbar button's label/icon to match
        the active data source mode."""
        if self.data_mode == "LCMS":
            self.gps_load_btn.config(text="📡  Load GPS (LCMS TXT Folder)")
        else:
            self.gps_load_btn.config(text="🌍  Load GPS (LiDAR KML/KMZ)")

    def _gps_load_dispatch(self):
        """Route the GPS-load button click to the correct loader for the
        currently active data source mode."""
        if self.data_mode == "LCMS":
            self.load_gps_lcms_txt_folder()
        else:
            self.load_gps_kml()

    def on_data_mode_changed(self, event=None):
        """Called when the user switches the Data Source dropdown between
        LIDAR and LCMS without restarting the app."""
        new_mode = self.mode_var.get()
        if new_mode == self.data_mode:
            return
        self.data_mode = new_mode
        self.root.title(f"PGR Survey Viewer v7  ·  {self.data_mode} Mode")
        self._refresh_gps_button()
        self._set_status(f"Data source switched to {self.data_mode} mode", color=ACCENT)

    def load_gps_lcms_txt_folder(self):
        """LCMS-mode GPS loader.

        LCMS surveys deliver GPS as GPS_Raw*.txt export files (one JSON
        object per line, each containing an "OdoDataRecord" + "NmeaLine")
        rather than the KML/KMZ trajectory files used for LiDAR surveys.

        This currently reuses the built-in GPS_Raw*.txt parsing pipeline
        (extract_gps_by_run_from_raw_txt_folder / parse_gps_raw_txt_file),
        which already targets this exact file format.

        >>> INTEGRATION POINT: once Samir's dedicated LCMS GPS-extraction
        script is provided, drop its parsing logic in here (or have it
        populate the same `runs: dict[run_number, list[coord_dict]]`
        shape used below) so it feeds the same gps_by_run pipeline that
        drives map drawing, frame sync, Save GPS, and Excel/CSV export. <<<
        """
        self.load_gps_raw_txt_folder()

    def load_gps_kml(self):
        if not HAVE_MAP:
            messagebox.showwarning("Map Disabled", "tkintermapview not installed.\npip install tkintermapview")
            return
        
        paths = filedialog.askopenfilenames(
            title="Select GPS KML/KMZ File(s)",
            filetypes=[("GPS Files", "*.kml *.kmz"), ("KML", "*.kml"), ("KMZ", "*.kmz"), ("All Files", "*.*")]
        )
        if not paths:
            return
        
        self._set_status("Loading GPS data...", color=WARNING)
        self.root.update()
        
        def worker():
            try:
                runs = extract_multiple_gps_by_run(list(paths))

                # Decimate any individual run that's too large
                total_before = sum(len(v) for v in runs.values())
                if total_before > GPS_DECIMATION_FACTOR:
                    # Decimate proportionally per-run so the per-run shape is preserved
                    for run_number in list(runs.keys()):
                        pts = runs[run_number]
                        target = max(2, int(len(pts) * GPS_DECIMATION_FACTOR / max(1, total_before)))
                        if len(pts) > target:
                            runs[run_number] = decimate_gps_points(pts, target)
                    total_after = sum(len(v) for v in runs.values())
                    decimated_msg = f" (decimated from {total_before} to {total_after})"
                else:
                    decimated_msg = ""

                self.root.after(0, self._gps_load_done, runs, list(paths), decimated_msg)
            except Exception as e:
                self.root.after(0, self._gps_load_error, str(e))
        
        threading.Thread(target=worker, daemon=True).start()
    
    def _gps_load_done(self, runs, paths, decimated_msg):
        if not runs:
            messagebox.showwarning("No GPS", "No valid coordinates found in selected files.")
            return

        self.gps_by_run = runs

        # Flat (legacy) view = all runs concatenated in run-number order (used by exports)
        all_coords = []
        for run_number in sorted(runs.keys(), key=lambda x: (x is None, x)):
            all_coords.extend(runs[run_number])
        self.gps_full_data = all_coords
        self.gps_coords = [(c["lat"], c["lon"]) for c in all_coords]
        self.current_gps_files = paths

        total_points = len(self.gps_coords)

        # Refresh whatever selection is currently active (single file / merged set)
        self._refresh_active_gps_selection()

        file_names = [Path(p).name for p in paths]
        label_txt = ", ".join(file_names[:2]) + ("…" if len(file_names) > 2 else "")
        n_runs = len([k for k in runs.keys() if k is not None])
        self._set_status(f"GPS loaded: {total_points} points across {n_runs} runs{decimated_msg}", color=SUCCESS)
    
    def _gps_load_error(self, error):
        messagebox.showerror("GPS Error", error)

    def load_gps_raw_txt_folder(self):
        """Embedded version of CombineTextFiles1.py: pick a root folder,
        recursively find GPS_Raw*.txt files, parse the NMEA/Odo JSON lines
        inside them, and feed the resulting points into the exact same
        gps_by_run pipeline used by the KML loader (so map drawing, frame
        sync, Save GPS, Export Excel/CSV all keep working unmodified)."""
        if not HAVE_MAP:
            messagebox.showwarning("Map Disabled", "tkintermapview not installed.\npip install tkintermapview")
            return

        root_folder = filedialog.askdirectory(title="Select Root Folder Containing GPS_Raw*.txt Files")
        if not root_folder:
            return

        self._set_status("Scanning for GPS_Raw*.txt files...", color=WARNING)
        self.root.update()

        def progress_cb(done, total, points_so_far):
            self.root.after(0, lambda: self._set_status(
                f"Parsing GPS_Raw files... {done}/{total} files, {points_so_far} points", color=WARNING))

        def worker():
            try:
                runs = extract_gps_by_run_from_raw_txt_folder(root_folder, progress_cb=progress_cb)

                total_before = sum(len(v) for v in runs.values())
                if total_before > GPS_DECIMATION_FACTOR:
                    for run_number in list(runs.keys()):
                        pts = runs[run_number]
                        target = max(2, int(len(pts) * GPS_DECIMATION_FACTOR / max(1, total_before)))
                        if len(pts) > target:
                            runs[run_number] = decimate_gps_points(pts, target)
                    total_after = sum(len(v) for v in runs.values())
                    decimated_msg = f" (decimated from {total_before} to {total_after})"
                else:
                    decimated_msg = ""

                self.root.after(0, self._raw_gps_load_done, runs, root_folder, decimated_msg)
            except Exception as e:
                self.root.after(0, self._gps_load_error, str(e))

        threading.Thread(target=worker, daemon=True).start()

    def _raw_gps_load_done(self, runs, root_folder, decimated_msg):
        if not runs:
            messagebox.showwarning("No GPS", "No valid GPS_Raw*.txt records found under the selected folder.")
            return

        # Merge into whatever GPS data is already loaded (KML and Raw TXT
        # runs can coexist; if a run number collides, raw points are appended).
        for run_number, pts in runs.items():
            self.gps_by_run.setdefault(run_number, []).extend(pts)

        all_coords = []
        for run_number in sorted(self.gps_by_run.keys(), key=lambda x: (x is None, x)):
            all_coords.extend(self.gps_by_run[run_number])
        self.gps_full_data = all_coords
        self.gps_coords = [(c["lat"], c["lon"]) for c in all_coords]
        self.current_gps_files = [root_folder]
        self.last_raw_gps_runs = runs  # kept around for the CSV export helper

        total_points = sum(len(v) for v in runs.values())

        self._refresh_active_gps_selection()

        n_runs = len([k for k in runs.keys() if k is not None])
        self._set_status(
            f"Raw GPS loaded: {total_points} points across {n_runs} runs from {Path(root_folder).name}{decimated_msg}",
            color=SUCCESS)

        # Offer to export the same CSV format CombineTextFiles1.py used to produce
        if messagebox.askyesno("Export CSV?", "Also save a combined CSV of this GPS_Raw data (same format as CombineTextFiles1.py)?"):
            out_path = filedialog.asksaveasfilename(
                title="Save Combined GPS CSV",
                defaultextension=".csv",
                filetypes=[("CSV Files", "*.csv"), ("All Files", "*.*")],
                initialfile="combined_gps_raw.csv"
            )
            if out_path:
                try:
                    export_raw_gps_runs_to_csv(self.gps_by_run, out_path)
                    messagebox.showinfo("Saved", f"GPS CSV saved to:\n{out_path}")
                except Exception as e:
                    messagebox.showerror("Export Error", str(e))

    def _get_active_pgr_run_numbers(self):
        """Return the list of run numbers (in display order) corresponding to
        whichever PGR file(s) are currently active — single-file mode or merged mode."""
        run_numbers = []
        if self.merged_mode:
            for entry in self.merged_index.file_entries:
                run_numbers.append(extract_run_number_from_pgr_filename(entry["path"].name))
        elif self.pgr_files and 0 <= self.current_file_idx < len(self.pgr_files):
            run_numbers.append(extract_run_number_from_pgr_filename(self.pgr_files[self.current_file_idx].name))
        return run_numbers

    def _refresh_active_gps_selection(self):
        """Recompute active_gps_coords / active_gps_full_data based on which
        PGR file(s) are currently loaded/selected, then rebuild the frame map
        and redraw the map. This is THE fix for 'marker moves across the
        whole GPS path when only one PGR file is selected'."""
        if not self.gps_by_run:
            self.active_gps_coords = []
            self.active_gps_full_data = []
            self.active_run_label = ""
            self.gps_frame_map = {}
            return

        run_numbers = self._get_active_pgr_run_numbers()
        # Drop unmatched (None) entries — they mean the filename didn't carry a run number
        matched_runs = [r for r in run_numbers if r is not None and r in self.gps_by_run]

        if matched_runs:
            active_full = []
            for r in matched_runs:
                active_full.extend(self.gps_by_run[r])
            if len(matched_runs) == 1:
                self.active_run_label = f"Run {matched_runs[0]:04d}"
            else:
                self.active_run_label = f"{len(matched_runs)} runs: " + ", ".join(f"{r:04d}" for r in matched_runs)
        else:
            # Fallback: no run-number match found (e.g. unrecognized filename pattern) ->
            # show everything rather than nothing, but say so clearly in the UI.
            active_full = list(self.gps_full_data)
            self.active_run_label = "All runs (no per-file match found)"

        self.active_gps_full_data = active_full
        self.active_gps_coords = [(c["lat"], c["lon"]) for c in active_full]

        self._rebuild_gps_frame_mapping()
        self._update_map_display()
        self.lbl_gps_file.config(text=f"GPS: {self.active_run_label} ({len(self.active_gps_coords)} pts)")

    def _rebuild_gps_frame_mapping(self):
        if not self.active_gps_coords or self.total_frames <= 1:
            self.gps_frame_map = {}
            return
        n = len(self.active_gps_coords)
        self.gps_frame_map = {}
        for i, coord in enumerate(self.active_gps_coords):
            fi = int(i * (self.total_frames - 1) / max(1, n - 1))
            self.gps_frame_map[fi] = coord
    
    def _update_map_display(self):
        if not HAVE_MAP or not self.map_widget or not self.active_gps_coords:
            return
        
        for path_obj in self.map_paths:
            try:
                path_obj.delete()
            except Exception:
                pass
        self.map_paths = []
        
        if self.car_marker:
            try:
                self.car_marker.delete()
            except Exception:
                pass
            self.car_marker = None
        
        # Build segments (optimized, smoothed) from the ACTIVE selection only
        segments = build_gps_segments_optimized(self.active_gps_coords)
        
        # Limit segments for performance
        max_segments = 100
        if len(segments) > max_segments:
            segments = segments[:max_segments]
        
        for seg in segments:
            if len(seg) >= 2:
                path_obj = self.map_widget.set_path(seg, color=ROUTE_COLOR, width=3)
                self.map_paths.append(path_obj)
        
        lat0, lon0 = self.active_gps_coords[0]
        self.map_widget.set_position(lat0, lon0)
        self.map_widget.set_zoom(17)
        
        self.car_marker = self.map_widget.set_marker(
            lat0, lon0, text="🚗",
            marker_color_circle=CAR_MARKER_COLOR,
            marker_color_outside="#880000"
        )
    
    def _update_gps_marker(self):
        if not self.active_gps_coords or not self.car_marker or self.total_frames <= 1:
            return
        
        if self.gps_frame_map:
            closest_frame = min(self.gps_frame_map.keys(), key=lambda f: abs(f - self.current_frame_idx))
            lat, lon = self.gps_frame_map[closest_frame]
        else:
            progress = self.current_frame_idx / (self.total_frames - 1)
            idx = int(progress * (len(self.active_gps_coords) - 1))
            lat, lon = self.active_gps_coords[idx]
        
        self.car_marker.set_position(lat, lon)
        
        if self.is_playing:
            if self.current_frame_idx % 5 == 0:
                self.map_widget.set_position(lat, lon)
        
        alt_text = ""
        if self.gps_frame_map:
            closest = min(self.gps_frame_map.keys(), key=lambda f: abs(f - self.current_frame_idx))
            gi = round(closest * (len(self.active_gps_coords) - 1) / max(1, self.total_frames - 1))
        else:
            gi = int((self.current_frame_idx / (self.total_frames - 1)) * (len(self.active_gps_coords) - 1))
        
        if 0 <= gi < len(self.active_gps_full_data):
            alt = self.active_gps_full_data[gi].get("altitude_m", 0)
            if alt > 0:
                alt_text = f"  Alt: {alt:.1f}m"
        
        self.lbl_gps_coord.config(text=f"GPS: {lat:.6f}, {lon:.6f}{alt_text}")
    
    def sync_video_from_map(self, coords):
        if not self.active_gps_coords or self.total_frames == 0:
            messagebox.showinfo("Info", "Load PGR and GPS files first.")
            return
        lat, lon = coords
        self._jump_video_to_gps(lat, lon)
    
    def _jump_video_to_gps(self, lat: float, lon: float):
        best_idx = min(range(len(self.active_gps_coords)), key=lambda i: math.hypot(self.active_gps_coords[i][0] - lat, self.active_gps_coords[i][1] - lon))
        progress = best_idx / max(1, len(self.active_gps_coords) - 1)
        target_frame = int(progress * (self.total_frames - 1))
        
        if self.is_playing:
            self.toggle_play()
        
        self.current_frame_idx = target_frame
        self.show_frame()
        self.map_widget.set_position(self.active_gps_coords[best_idx][0], self.active_gps_coords[best_idx][1])
        self._set_status(f"Jumped to GPS #{best_idx} → Frame {target_frame}", color=ACCENT)
    
    def on_map_type_changed(self, _event=None):
        if HAVE_MAP and self.map_widget:
            tile_url = MAP_TILES.get(self.map_type_var.get())
            if tile_url:
                self.map_widget.set_tile_server(tile_url, max_zoom=22)
    
    # ── Save GPS ──────────────────────────────────────────────
    def save_gps_to_files(self):
        if not self.gps_coords:
            messagebox.showwarning("Save GPS", "No GPS data loaded.")
            return
        out_dir = filedialog.askdirectory(title="Select output folder for GPS files")
        if not out_dir:
            return
        out_dir = Path(out_dir)
        out_dir.mkdir(parents=True, exist_ok=True)
        
        frame_map = {}
        if self.total_frames > 0:
            for i in range(len(self.gps_coords)):
                fi = int(i * (self.total_frames - 1) / max(1, len(self.gps_coords) - 1))
                frame_map[i] = fi
        
        csv_path = out_dir / "GPS_Export.csv"
        kml_path = out_dir / "GPS_Export.kml"
        n_csv = save_gps_to_csv(self.gps_full_data, csv_path, frame_map)
        n_kml = save_gps_to_kml(self.gps_full_data, kml_path, "PGR_Survey_GPS_Track")
        
        messagebox.showinfo("GPS Export", f"Saved to: {out_dir}\n\nCSV: {n_csv} records\nKML: {n_kml} points (line only)")
    
    # ── Excel export (streaming, memory efficient) ──────────────
    def export_to_excel(self):
        if not HAVE_PANDAS:
            messagebox.showwarning("Export", "pandas and openpyxl required.\npip install pandas openpyxl")
            return
        if not self.pgr_files:
            messagebox.showwarning("Export", "Load a PGR folder first.")
            return
        if not self.gps_coords:
            messagebox.showwarning("Export", "Load a GPS KML/KMZ file first.")
            return
        
        save_path = filedialog.asksaveasfilename(
            title="Save Survey Export",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile="PGR_Survey_Export.xlsx"
        )
        if not save_path:
            return
        
        self._set_status("Exporting to Excel... (streaming)", color=WARNING)
        self.root.update()
        
        def worker():
            try:
                # Use streaming exporter
                exporter = StreamingExcelExporter(Path(save_path))
                exporter.start()
                
                cum = [0.0]
                for i in range(1, len(self.gps_coords)):
                    cum.append(cum[-1] + _haversine_m(*self.gps_coords[i-1], *self.gps_coords[i]))
                
                n = len(self.gps_coords)
                row_count = 0
                
                if self.merged_mode and self.merged_index.total_frames > 0:
                    for fe in self.merged_index.file_entries:
                        frame_count = len(fe["frames"])
                        if frame_count == 0:
                            continue
                        fname = fe["path"].name
                        for fi in range(frame_count):
                            p = fi / max(1, frame_count - 1)
                            gi = int(p * (n - 1))
                            lat, lon = self.gps_coords[gi]
                            alt = self.gps_full_data[gi].get("altitude_m", 0) if gi < len(self.gps_full_data) else 0
                            name = self.gps_full_data[gi].get("name", "") if gi < len(self.gps_full_data) else ""
                            exporter.add_row({
                                "File": fname, "Frame": fi, "Latitude": lat,
                                "Longitude": lon, "Altitude_m": alt, "Point_Name": name,
                                "Chainage_m": round(cum[gi], 2)
                            })
                            row_count += 1
                else:
                    for pgr_path in self.pgr_files:
                        try:
                            mm, frames = scan_pgr_frames_optimized(pgr_path)
                            frame_count = len(frames)
                            mm.close()
                        except Exception:
                            continue
                        fname = pgr_path.name
                        for fi in range(frame_count):
                            p = fi / max(1, frame_count - 1)
                            gi = int(p * (n - 1))
                            lat, lon = self.gps_coords[gi]
                            alt = self.gps_full_data[gi].get("altitude_m", 0) if gi < len(self.gps_full_data) else 0
                            name = self.gps_full_data[gi].get("name", "") if gi < len(self.gps_full_data) else ""
                            exporter.add_row({
                                "File": fname, "Frame": fi, "Latitude": lat,
                                "Longitude": lon, "Altitude_m": alt, "Point_Name": name,
                                "Chainage_m": round(cum[gi], 2)
                            })
                            row_count += 1
                
                exporter.finish()
                self.root.after(0, self._export_done_excel, True, row_count, save_path)
            except Exception as e:
                self.root.after(0, self._export_done_excel, False, 0, str(e))
        
        threading.Thread(target=worker, daemon=True).start()
    
    def _export_done_excel(self, success, count, msg):
        if success:
            messagebox.showinfo("Export Complete", f"Saved {count} records:\n{msg}")
            self._set_status(f"Export complete: {count} records", color=SUCCESS)
        else:
            messagebox.showerror("Export Error", f"Failed to export:\n{msg}")
            self._set_status("Export failed.", color=DANGER)
    
    # ── Cam0 Images + Excel Export (batched, parallel) ─────────
    def export_cam0_and_excel(self):
        if not HAVE_PANDAS:
            messagebox.showwarning("Export Required", "pandas and openpyxl required.\npip install pandas openpyxl")
            return
        if self.total_frames == 0:
            messagebox.showwarning("Export Alert", "Please load a PGR folder first.")
            return
        if not self.gps_coords:
            messagebox.showwarning("Export Alert", "Please load a GPS KML/KMZ file first.")
            return
        
        out_dir = filedialog.askdirectory(title="Select Folder to Save Cam 0 Images & Excel Data")
        if not out_dir:
            return
        out_dir = Path(out_dir)
        
        self._set_status("Exporting Cam 0 & Excel Data in parallel...", color=WARNING)
        self.is_indexing = True
        self._show_progress(True)
        self._stop_playback_if_running()
        
        def worker():
            try:
                # Pre-calculate GPS chainage
                cum = [0.0]
                for i in range(1, len(self.gps_coords)):
                    cum.append(cum[-1] + _haversine_m(*self.gps_coords[i-1], *self.gps_coords[i]))
                
                # Create images directory
                images_dir = out_dir / "Cam0_Images"
                images_dir.mkdir(exist_ok=True)
                
                # Batch process frames
                rows = []
                n = len(self.gps_coords)
                
                for i in range(self.total_frames):
                    if self.merged_mode:
                        mm, frame_data = self.merged_index.get_frame_data(i)
                    else:
                        mm = self._single_mm
                        frame_data = self._single_frames[i]
                    
                    # Decode Camera 0
                    img = decode_camera_frame_optimized(mm, frame_data["planes"], 0, 
                                                         self.r_gain, self.g_gain, self.b_gain, 
                                                         True, decode_quality=90)
                    
                    if img:
                        img_name = f"Cam0_Frame_{i:06d}.jpg"
                        img_path = images_dir / img_name
                        img.save(img_path, "JPEG", quality=85, optimize=True)
                        
                        progress = i / max(1, self.total_frames - 1)
                        gi = int(progress * (n - 1))
                        lat, lon = self.gps_coords[gi]
                        alt = self.gps_full_data[gi].get("altitude_m", 0) if gi < len(self.gps_full_data) else 0
                        
                        rows.append({
                            "Image_Name": img_name,
                            "Frame": i,
                            "Latitude": lat,
                            "Longitude": lon,
                            "Altitude_m": alt,
                            "Chainage_m": round(cum[gi], 2)
                        })
                    
                    if i % 25 == 0:
                        pct = (i / self.total_frames) * 100.0
                        self.root.after(0, self._set_progress, pct, f"Exporting {i}/{self.total_frames} frames...")
                
                # Save Excel
                if rows:
                    df = pd.DataFrame(rows)
                    excel_path = out_dir / "Cam0_GPS_Export.xlsx"
                    with pd.ExcelWriter(excel_path, engine="openpyxl") as wr:
                        df.to_excel(wr, index=False, sheet_name="Cam0_Export")
                        ws = wr.sheets["Cam0_Export"]
                        for col in ws.columns:
                            max_len = max((len(str(c.value or "")) for c in col), default=10)
                            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)
                
                self.root.after(0, self._export_done, True, len(rows), str(out_dir))
                
            except Exception as e:
                self.root.after(0, self._export_done, False, 0, str(e))
        
        threading.Thread(target=worker, daemon=True).start()
    
    def _export_done(self, success, count, msg):
        self.is_indexing = False
        self._show_progress(False)
        if success:
            messagebox.showinfo("Export Complete", f"Successfully exported {count} Cam 0 images and GPS Excel sheet to:\n{msg}")
            self._set_status(f"Export complete: {count} images saved", color=SUCCESS)
        else:
            messagebox.showerror("Export Error", f"Failed to export:\n{msg}")
            self._set_status("Export failed.", color=DANGER)

    # ── Video export (choose quality, encode current view mode/camera) ──
    def export_video_dialog(self):
        """Open a small dialog to pick quality preset + frame range, then export."""
        if not HAVE_CV2:
            messagebox.showwarning("Video Export", "opencv-python required.\npip install opencv-python")
            return
        if (self.merged_mode and self.merged_index.total_frames == 0) or \
           (not self.merged_mode and not self._single_frames):
            messagebox.showwarning("Video Export", "Load PGR file(s) first.")
            return

        dlg = tk.Toplevel(self.root)
        dlg.title("Export Video")
        dlg.configure(bg=BG_PANEL)
        dlg.resizable(False, False)
        dlg.transient(self.root)
        dlg.grab_set()

        pad = {"padx": 14, "pady": 6}

        tk.Label(dlg, text="🎬 Export Video", bg=BG_PANEL, fg=FG_BRIGHT,
                 font=("Segoe UI", 12, "bold")).grid(row=0, column=0, columnspan=2, sticky="w", **pad)

        tk.Label(dlg, text="Quality:", bg=BG_PANEL, fg=FG_DIM, font=("Segoe UI", 9)).grid(row=1, column=0, sticky="w", **pad)
        quality_var = tk.StringVar(value="Medium (960p, balanced)")
        quality_combo = ttk.Combobox(dlg, textvariable=quality_var, values=list(VIDEO_QUALITY_PRESETS.keys()),
                                      state="readonly", width=26, font=("Segoe UI", 9))
        quality_combo.grid(row=1, column=1, sticky="w", **pad)

        tk.Label(dlg, text="FPS:", bg=BG_PANEL, fg=FG_DIM, font=("Segoe UI", 9)).grid(row=2, column=0, sticky="w", **pad)
        fps_var = tk.IntVar(value=max(1, self.fps_var.get()))
        tk.Spinbox(dlg, from_=1, to=60, width=6, textvariable=fps_var, font=("Courier New", 9),
                   bg=BG_CARD, fg=ACCENT, relief="flat").grid(row=2, column=1, sticky="w", **pad)

        tk.Label(dlg, text="View:", bg=BG_PANEL, fg=FG_DIM, font=("Segoe UI", 9)).grid(row=3, column=0, sticky="w", **pad)
        view_label = {"single": "Single Camera", "dual": "Dual Panoramic", "grid": "All 6 Cameras Grid"}.get(self.view_mode, "Single Camera")
        tk.Label(dlg, text=f"{view_label}  (uses current viewer settings)", bg=BG_PANEL, fg=FG_MAIN,
                 font=("Segoe UI", 9)).grid(row=3, column=1, sticky="w", **pad)

        tk.Label(dlg, text="Range:", bg=BG_PANEL, fg=FG_DIM, font=("Segoe UI", 9)).grid(row=4, column=0, sticky="w", **pad)
        range_var = tk.StringVar(value="all")
        range_frame = tk.Frame(dlg, bg=BG_PANEL)
        range_frame.grid(row=4, column=1, sticky="w", **pad)
        tk.Radiobutton(range_frame, text="All frames", variable=range_var, value="all",
                       bg=BG_PANEL, fg=FG_MAIN, selectcolor=BG_CARD, activebackground=BG_PANEL,
                       font=("Segoe UI", 9)).pack(anchor="w")
        tk.Radiobutton(range_frame, text="Current frame to end", variable=range_var, value="from_current",
                       bg=BG_PANEL, fg=FG_MAIN, selectcolor=BG_CARD, activebackground=BG_PANEL,
                       font=("Segoe UI", 9)).pack(anchor="w")

        btn_frame = tk.Frame(dlg, bg=BG_PANEL)
        btn_frame.grid(row=5, column=0, columnspan=2, pady=(10, 14))

        def on_export():
            preset_name = quality_var.get()
            fps = max(1, fps_var.get())
            start_frame = self.current_frame_idx if range_var.get() == "from_current" else 0
            dlg.destroy()
            self._start_video_export(preset_name, fps, start_frame)

        tk.Button(btn_frame, text="✓ Export", bg=BTN_PRIMARY, fg="white", font=("Segoe UI", 10, "bold"),
                  relief="flat", padx=16, pady=6, cursor="hand2", command=on_export).pack(side=tk.LEFT, padx=6)
        tk.Button(btn_frame, text="Cancel", bg=BG_CARD, fg=FG_DIM, font=("Segoe UI", 9),
                  relief="flat", padx=16, pady=6, cursor="hand2", command=dlg.destroy).pack(side=tk.LEFT, padx=6)

    def _start_video_export(self, preset_name: str, fps: int, start_frame: int):
        save_path = filedialog.asksaveasfilename(
            title="Save Video As",
            defaultextension=".mp4",
            filetypes=[("MP4 Video", "*.mp4")],
            initialfile="PGR_Survey_Export.mp4"
        )
        if not save_path:
            return

        preset = VIDEO_QUALITY_PRESETS.get(preset_name, VIDEO_QUALITY_PRESETS["Medium (960p, balanced)"])
        end_frame = self.total_frames - 1
        if end_frame < start_frame:
            messagebox.showwarning("Video Export", "Invalid frame range.")
            return

        self._stop_playback_if_running()
        self.is_indexing = True
        self._show_progress(True)
        self._set_progress(0, "Preparing video export...")
        self._set_status("Exporting video...", color=WARNING)

        # Snapshot current viewer settings so they don't change mid-export
        view_mode = self.view_mode
        cam_idx = self.current_cam_idx
        r_gain, g_gain, b_gain = self.r_gain, self.g_gain, self.b_gain
        merged_mode = self.merged_mode

        def get_frame_source(idx):
            if merged_mode:
                return self.merged_index.get_frame_data(idx)
            else:
                return self._single_mm, self._single_frames[idx]

        def decode_for_export(idx):
            mm, frame_data = get_frame_source(idx)
            if view_mode == "dual":
                return decode_dual_panoramic_optimized(mm, frame_data["planes"], r_gain, g_gain, b_gain, decode_quality=100)
            elif view_mode == "grid":
                return decode_all_cameras_grid_optimized(mm, frame_data["planes"], r_gain, g_gain, b_gain, decode_quality=100)
            else:
                return decode_camera_frame_optimized(mm, frame_data["planes"], cam_idx, r_gain, g_gain, b_gain, decode_quality=100)

        def worker():
            writer = None
            frames_written = 0
            try:
                total = end_frame - start_frame + 1

                # Decode the first frame to determine output dimensions
                first_img = decode_for_export(start_frame)
                base_w, base_h = first_img.size
                scale = preset["scale"]
                out_w = max(2, int(base_w * scale)) & ~1   # even dimensions required by most codecs
                out_h = max(2, int(base_h * scale)) & ~1

                fourcc = cv2.VideoWriter_fourcc(*preset["fourcc"])
                writer = cv2.VideoWriter(str(save_path), fourcc, fps, (out_w, out_h))
                if not writer.isOpened():
                    # Fallback to mp4v if the preferred codec isn't available on this system
                    fourcc = cv2.VideoWriter_fourcc(*"mp4v")
                    writer = cv2.VideoWriter(str(save_path), fourcc, fps, (out_w, out_h))
                if not writer.isOpened():
                    raise RuntimeError("Could not open video writer (codec unsupported on this system).")

                for idx in range(start_frame, end_frame + 1):
                    img = first_img if idx == start_frame else decode_for_export(idx)
                    if (out_w, out_h) != img.size:
                        img = img.resize((out_w, out_h), Image.BILINEAR)
                    frame_bgr = cv2.cvtColor(np.array(img), cv2.COLOR_RGB2BGR)
                    writer.write(frame_bgr)
                    frames_written += 1

                    if frames_written % 10 == 0:
                        pct = (frames_written / max(1, total)) * 100.0
                        self.root.after(0, self._set_progress, pct, f"Encoding frame {frames_written}/{total}...")

                writer.release()
                self.root.after(0, self._video_export_done, True, frames_written, str(save_path))
            except Exception as e:
                if writer is not None:
                    try:
                        writer.release()
                    except Exception:
                        pass
                self.root.after(0, self._video_export_done, False, frames_written, str(e))

        threading.Thread(target=worker, daemon=True).start()

    def _video_export_done(self, success, count, msg):
        self.is_indexing = False
        self._show_progress(False)
        if success:
            messagebox.showinfo("Video Export Complete", f"Saved {count} frames to:\n{msg}")
            self._set_status(f"Video export complete: {count} frames", color=SUCCESS)
        else:
            messagebox.showerror("Video Export Error", f"Failed to export video:\n{msg}")
            self._set_status("Video export failed.", color=DANGER)

    # ── Helpers ──────────────────────────────────────────────
    def _set_status(self, msg: str, color: str = FG_MAIN):
        self.lbl_status.config(text=f"⬤  {msg}", fg=color)
    
    def _progress_cb(self, pct: float, msg: str):
        self.root.after(0, lambda: self._set_progress(pct, msg))
    
    def _set_progress(self, pct: float, msg: str):
        self.progress_var.set(pct)
        self.lbl_progress.config(text=msg)


# Helper functions for GPS export
def save_gps_to_csv(coords_data: list, output_path: Path, frame_mapping: dict = None) -> int:
    if not coords_data:
        return 0
    fieldnames = ["index", "lat", "lon", "altitude_m", "name", "frame"]
    rows = []
    for i, coord in enumerate(coords_data):
        rows.append({
            "index": i,
            "lat": coord["lat"],
            "lon": coord["lon"],
            "altitude_m": coord.get("altitude_m", 0),
            "name": coord.get("name", ""),
            "frame": frame_mapping.get(i, -1) if frame_mapping else -1,
        })
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    return len(rows)


def save_gps_to_kml(coords_data: list, output_path: Path, name: str = "GPS Track") -> int:
    valid = [c for c in coords_data if c.get("lat") is not None]
    if not valid:
        return 0
    
    def esc(s):
        return str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    
    lines = [
        '<?xml version="1.0" encoding="UTF-8"?>',
        '<kml xmlns="http://www.opengis.net/kml/2.2">',
        '<Document>',
        f'  <name>{esc(name)}</name>',
        '  <Style id="trackLine">',
        '    <LineStyle><color>ff00d4ff</color><width>3</width></LineStyle>',
        '    <PolyStyle><fill>0</fill></PolyStyle>',
        '  </Style>',
        '  <Placemark>',
        f'    <name>{esc(name)} Track</name>',
        '    <styleUrl>#trackLine</styleUrl>',
        '    <LineString>',
        '      <tessellate>1</tessellate>',
        '      <coordinates>',
    ]
    for c in valid:
        lines.append(f'        {c["lon"]},{c["lat"]},{c.get("altitude_m",0)}')
    lines += [
        '      </coordinates>',
        '    </LineString>',
        '  </Placemark>',
        '</Document>', '</kml>',
    ]
    output_path.write_text("\n".join(lines), encoding="utf-8")
    return len(valid)


# ─────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import io
    
    root = tk.Tk()
    root.withdraw()
    
    if not HAVE_CODECS:
        print("\n" + "="*60)
        print("WARNING: imagecodecs not installed!")
        print("PGR file decoding will not work.")
        print("Install with: pip install imagecodecs")
        print("="*60 + "\n")
    
    if not HAVE_MAP:
        print("\n" + "="*60)
        print("WARNING: tkintermapview not installed!")
        print("Map display will be disabled.")
        print("Install with: pip install tkintermapview")
        print("="*60 + "\n")

    if not HAVE_CV2:
        print("\n" + "="*60)
        print("WARNING: opencv-python not installed!")
        print("Video export will be disabled.")
        print("Install with: pip install opencv-python")
        print("="*60 + "\n")

    # First screen: choose data source (LiDAR -> KML/KMZ GPS, LCMS -> TXT GPS),
    # then open the PGR viewer configured for that mode.
    chosen_mode = DataModeDialog.ask(root)
    if not chosen_mode:
        root.destroy()
        sys.exit(0)

    root.deiconify()
    app = PGRViewerPro(root, data_mode=chosen_mode)
    root.mainloop()
